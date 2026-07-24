import json
import csv
import re
from io import StringIO
from datetime import datetime
import aiosqlite
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, InlineKeyboardButton, InputMediaPhoto
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup

from config import ADMIN_IDS, BOT_TOKEN
from database import *
from keyboards import *

router = Router()


# ============================================================
# 📌 СОСТОЯНИЯ (FSM)
# ============================================================

class ApplicationForm(StatesGroup):
    waiting_for_answers = State()
    waiting_photo_old = State()
    waiting_photo_new = State()
    waiting_contact_message = State()
    waiting_test_answers = State()


class RoleForm(StatesGroup):
    waiting_clan_id = State()
    waiting_user_id = State()
    waiting_username = State()
    waiting_name = State()
    waiting_role_type = State()


# ============================================================
# 🏠 СТАРТ
# ============================================================

@router.message(Command('start'))
async def cmd_start(message: Message):
    clan = await get_clan_by_user(message.from_user.id)
    if clan:
        await message.answer('🏠 Добро пожаловать в KLAN KAIF!\n\nВыберите действие:', reply_markup=leader_menu())
    else:
        await message.answer('🏠 Добро пожаловать в KLAN KAIF!\n\nВыберите действие:', reply_markup=main_menu())


# ============================================================
# ⚙️ АДМИН
# ============================================================

@router.message(Command('admin'))
async def cmd_admin(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        await message.answer('⛔ У вас нет прав администратора.')
        return
    await message.answer('⚙️ АДМИН-ПАНЕЛЬ KLAN KAIF\n\nВыберите действие:', reply_markup=admin_menu())


# ============================================================
# 🔙 ШАГ НАЗАД
# ============================================================

@router.callback_query(F.data == 'back_to_main')
async def back_to_main(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    
    # Проверяем, был ли тестовый режим
    data = await state.get_data()
    is_test = data.get('is_test_mode', False)
    
    if is_test:
        # Если тестовый режим — показываем меню кандидата
        await callback.message.edit_text(
            '🏠 Добро пожаловать в KLAN KAIF!\n\nВыберите действие:',
            reply_markup=main_menu()
        )
        return
    
    await state.clear()
    clan = await get_clan_by_user(callback.from_user.id)
    if clan:
        await callback.message.edit_text('🏠 Главное меню:', reply_markup=leader_menu())
    else:
        await callback.message.edit_text('🏠 Главное меню:', reply_markup=main_menu())


# ============================================================
# 🧪 ТЕСТОВАЯ АНКЕТА (ДЛЯ АДМИНОВ) - СТАРАЯ ВЕРСИЯ
# ============================================================

@router.callback_query(F.data == 'admin_test_application')
async def admin_test_application(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    await callback.message.edit_text(
        '🧪 ТЕСТОВАЯ АНКЕТА\n\n'
        'Нажмите "Написать тестовую анкету", чтобы отправить заявку как кандидат.\n\n'
        '📌 Анкета будет выглядеть как обычная заявка, но с пометкой "🧪 ТЕСТ"',
        reply_markup=test_application_menu()
    )


@router.callback_query(F.data == 'write_test_application')
async def write_test_application(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    await state.set_state(ApplicationForm.waiting_test_answers)

    await callback.message.edit_text(
        '📝 НАПИШИТЕ ТЕСТОВУЮ АНКЕТУ\n\n'
        '📋 Скопируйте шаблон и заполните:\n'
        '━━━━━━━━━━━━━━━━━━━━━━\n'
        'Имя: \n'
        'Возраст: \n'
        'Ник: \n'
        'ID: \n'
        'Часовой пояс (МСК): \n'
        '━━━━━━━━━━━━━━━━━━━━━━\n\n'
        '📌 ПРИМЕР:\n'
        'Имя: Тестовый Пользователь\n'
        'Возраст: 25\n'
        'Ник: Test_User\n'
        'ID: 123456789\n'
        'Часовой пояс (МСК): +0\n\n'
        '⚠️ После заполнения выберите клан для отправки.',
        reply_markup=back_button('back_to_test')
    )


@router.message(ApplicationForm.waiting_test_answers)
async def receive_test_application(message: Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        await message.answer('⛔ У вас нет прав')
        await state.clear()
        return

    text = message.text.strip()
    lines = [line.strip() for line in text.split('\n') if line.strip()]

    cleaned_lines = []
    for line in lines:
        if ':' in line or '：' in line:
            parts = re.split(r'[:：]', line, 1)
            if len(parts) == 2:
                cleaned_lines.append(parts[1].strip())
            else:
                cleaned_lines.append(line.strip())
        else:
            cleaned_lines.append(line.strip())

    if len(cleaned_lines) != 5:
        await message.answer(
            '❌ Нужно ровно 5 полей!\n\n'
            '📌 ПРАВИЛЬНЫЙ ФОРМАТ:\n'
            'Имя: Тестовый Пользователь\n'
            'Возраст: 25\n'
            'Ник: Test_User\n'
            'ID: 123456789\n'
            'Часовой пояс (МСК): +0',
            reply_markup=back_button('back_to_test')
        )
        return

    answers = {
        'name': cleaned_lines[0],
        'age': cleaned_lines[1],
        'nickname': cleaned_lines[2],
        'id': cleaned_lines[3],
        'timezone': cleaned_lines[4]
    }

    try:
        age = int(answers['age'])
        if age < 10 or age > 99:
            raise ValueError
    except:
        await message.answer(
            '❌ Возраст должен быть числом от 10 до 99!\n'
            'Попробуйте снова:',
            reply_markup=back_button('back_to_test')
        )
        return

    try:
        tz = int(answers['timezone'])
        if tz < -12 or tz > 12:
            raise ValueError
    except:
        await message.answer(
            '❌ Часовой пояс должен быть числом от -12 до +12!\n'
            'Попробуйте снова:',
            reply_markup=back_button('back_to_test')
        )
        return

    await state.update_data(test_answers=answers)
    await state.set_state(None)

    await message.answer(
        '✅ Анкета сохранена!\n\n'
        'Теперь выберите клан для отправки тестовой заявки:',
        reply_markup=clan_choice_for_test()
    )


@router.callback_query(F.data.startswith('test_clan_'))
async def test_select_clan(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    clan_id = int(callback.data.split('_')[2])
    clan = await get_clan(clan_id)
    if not clan:
        await callback.message.answer('❌ Клан не найден')
        return

    data = await state.get_data()
    answers = data.get('test_answers', {})

    if not answers:
        await callback.message.answer('❌ Анкета не найдена. Попробуйте снова.')
        return

    app_id = await add_application(
        callback.from_user.id,
        'test_user',
        clan_id,
        answers
    )

    clan_id, name, leader_id, leader_username, leader_name, deputy_id, deputy_username, deputy_name, _ = clan

    text = (
        f'🧪 ТЕСТОВАЯ ЗАЯВКА #{app_id} В КЛАН {name}\n\n'
        f'От: @{callback.from_user.username or "test"} (ID: {callback.from_user.id})\n'
        f'Дата: {datetime.now().strftime("%d.%m.%Y, %H:%M")}\n\n'
        f'📝 АНКЕТА:\n'
        f'1. Имя: {answers.get("name", "")}\n'
        f'2. Возраст: {answers.get("age", "")}\n'
        f'3. Ник: {answers.get("nickname", "")}\n'
        f'4. ID: {answers.get("id", "")}\n'
        f'5. Часовой пояс (МСК): {answers.get("timezone", "")}\n\n'
        f'📸 Скрины: [тестовые]'
    )

    # Отправка и лидеру, и заму
    if leader_id:
        try:
            await callback.bot.send_message(
                leader_id,
                text,
                reply_markup=review_buttons(app_id)
            )
        except Exception as e:
            print(f"Ошибка отправки лидеру: {e}")

    if deputy_id:
        try:
            await callback.bot.send_message(
                deputy_id,
                text,
                reply_markup=review_buttons(app_id)
            )
        except Exception as e:
            print(f"Ошибка отправки заму: {e}")

    await state.clear()

    await callback.message.edit_text(
        f'✅ Тестовая заявка #{app_id} создана!\n'
        f'📨 Отправлена лидеру и заму клана {name}.\n\n'
        f'🧪 Это тестовая заявка — она помечена как "ТЕСТ" в базе данных.',
        reply_markup=admin_menu()
    )


# ============================================================
# 🧑‍💻 СТАТЬ КАНДИДАТОМ (ДЛЯ АДМИНОВ) - НОВАЯ ВЕРСИЯ
# ============================================================

@router.callback_query(F.data == 'admin_become_candidate')
async def admin_become_candidate(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    # Показываем ТОЧНО ТАКОЕ ЖЕ МЕНЮ, как у кандидата
    await callback.message.edit_text(
        '🏠 Добро пожаловать в KLAN KAIF!\n\n'
        'Выберите действие:',
        reply_markup=main_menu()
    )
    
    # Сохраняем в state что это тестовый режим
    await state.update_data(is_test_mode=True)


# ============================================================
# 🔙 ВЫХОД ИЗ ТЕСТОВОГО РЕЖИМА
# ============================================================

@router.callback_query(F.data == 'exit_test_mode')
async def exit_test_mode(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    await state.clear()
    
    # Возвращаемся в админ-панель
    await callback.message.edit_text(
        '⚙️ АДМИН-ПАНЕЛЬ KLAN KAIF\n\nВыберите действие:',
        reply_markup=admin_menu()
    )


# ============================================================
# 👥 ЧЁРНЫЙ СПИСОК
# ============================================================

@router.message(Command('blacklist'))
async def cmd_blacklist(message: Message):
    clan = await get_clan_by_user(message.from_user.id)
    if not clan:
        await message.answer('⛔ Только лидеры и замы могут управлять чёрным списком.')
        return

    args = message.text.split()
    if len(args) < 2:
        await message.answer(
            '📖 Команды:\n/blacklist add @username причина — добавить в ЧС\n/blacklist remove @username — удалить из ЧС\n/blacklist list — показать список')
        return

    action = args[1]

    if action == 'list':
        blacklist = await get_blacklist()
        if not blacklist:
            await message.answer('👥 Чёрный список пуст.')
            return
        text = '👥 ЧЁРНЫЙ СПИСОК:\n\n'
        for item in blacklist:
            text += f'ID: {item[1]}\nПричина: {item[2]}\nДобавлен: {item[4][:10] if item[4] else "неизвестно"}\n\n'
        await message.answer(text)
        return

    if len(args) < 3:
        await message.answer('❌ Укажите пользователя. Пример: /blacklist add @username причина')
        return

    username = args[2].replace('@', '')
    reason = ' '.join(args[3:]) if len(args) > 3 else 'Не указана'

    async with aiosqlite.connect(DB_PATH) as db:
        if action == 'add':
            async with db.execute('SELECT DISTINCT user_id FROM applications WHERE username = ?',
                                  (username,)) as cursor:
                rows = await cursor.fetchall()
                if rows:
                    user_id = rows[0][0]
                    await add_to_blacklist(user_id, reason, message.from_user.id)
                    await message.answer(f'✅ @{username} добавлен в чёрный список.\nПричина: {reason}')
                else:
                    await message.answer(f'❌ Пользователь @{username} не найден в заявках.')
        elif action == 'remove':
            async with db.execute('SELECT DISTINCT user_id FROM applications WHERE username = ?',
                                  (username,)) as cursor:
                rows = await cursor.fetchall()
                if rows:
                    user_id = rows[0][0]
                    await remove_from_blacklist(user_id)
                    await message.answer(f'✅ @{username} удалён из чёрного списка.')
                else:
                    await message.answer(f'❌ Пользователь @{username} не найден.')


# ============================================================
# 📋 КОПИРОВАТЬ ШАБЛОН
# ============================================================

@router.callback_query(F.data == 'copy_template')
async def copy_template(callback: CallbackQuery):
    await callback.answer()

    template = (
        'Имя: \n'
        'Возраст: \n'
        'Ник: \n'
        'ID: \n'
        'Часовой пояс (МСК): '
    )

    await callback.message.answer(template)


# ============================================================
# ⏭️ ПРОПУСТИТЬ ФОТО 2
# ============================================================

@router.callback_query(F.data == 'skip_photo')
async def skip_photo(callback: CallbackQuery, state: FSMContext):
    await callback.answer()

    data = await state.get_data()
    app_id = data.get('app_id')
    clan_name = data.get('clan_name')
    answers = data.get('answers', {})
    photo_old = data.get('photo_old')
    is_test = data.get('is_test_mode', False)

    if not app_id:
        await callback.message.answer('Ошибка. Попробуйте начать заново через /start')
        await state.clear()
        return

    await state.clear()

    try:
        clan = await get_clan_by_name(clan_name)
        if not clan:
            await callback.message.answer('❌ Клан не найден.')
            return

        clan_id, name, leader_id, leader_username, leader_name, deputy_id, deputy_username, deputy_name, _ = clan

        text = (
            f'🔔 НОВАЯ ЗАЯВКА #{app_id} В КЛАН {clan_name}\n\n'
            f'От: @{callback.from_user.username or "unknown"} (ID: {callback.from_user.id})\n'
            f'Дата: {datetime.now().strftime("%d.%m.%Y, %H:%M")}\n\n'
            f'📝 АНКЕТА:\n'
            f'1. Имя: {answers.get("name", "")}\n'
            f'2. Возраст: {answers.get("age", "")}\n'
            f'3. Ник: {answers.get("nickname", "")}\n'
            f'4. ID: {answers.get("id", "")}\n'
            f'5. Часовой пояс (МСК): {answers.get("timezone", "")}\n\n'
            f'📸 Скринов: 1 (второе фото пропущено)'
        )

        # ОБНОВЛЯЕМ КОЛИЧЕСТВО ФОТО
        await update_application_has_photos(app_id, 1)

        # Отправка лидеру
        if leader_id:
            try:
                await callback.bot.send_photo(
                    leader_id,
                    photo=photo_old,
                    caption=text,
                    reply_markup=review_buttons(app_id)
                )
            except Exception as e:
                print(f"Ошибка отправки лидеру: {e}")

        # Отправка заму
        if deputy_id:
            try:
                await callback.bot.send_photo(
                    deputy_id,
                    photo=photo_old,
                    caption=text,
                    reply_markup=review_buttons(app_id)
                )
            except Exception as e:
                print(f"Ошибка отправки заму: {e}")

        # Если это тестовый режим, показываем кнопку выхода
        if is_test:
            await callback.message.edit_text(
                f'⏭️ Вы пропустили второе фото.\n'
                f'✅ ТЕСТОВАЯ заявка #{app_id} отправлена на рассмотрение!\n'
                f'Ожидайте решения лидера или зама.',
                reply_markup=exit_test_button()
            )
        else:
            await callback.message.edit_text(
                f'⏭️ Вы пропустили второе фото.\n'
                f'✅ Заявка #{app_id} отправлена на рассмотрение!\n'
                f'Ожидайте решения лидера или зама.',
                reply_markup=after_apply_buttons()
            )

    except Exception as e:
        await callback.message.answer(f'❌ Ошибка: {e}\nЗаявка сохранена.')


# ============================================================
# ℹ️ О КЛАНАХ
# ============================================================

@router.callback_query(F.data == 'about_clans')
async def about_clans(callback: CallbackQuery):
    await callback.answer()
    clans = await get_clans()
    text = '🏆 KLAN KAIF:\n\n'
    text += '📌 ОБЩИЕ ТРЕБОВАНИЯ ДЛЯ ВСЕХ:\n• Адекватность\n• Актив в беседе и игре\n• Смена ника с припиской KAIF\n• Участие в мероприятиях клана\n• Для новичков — неделя на показ активности\n\n'

    clan_data = {
        1: {'emoji': '🔴', 'name': 'KAIF — основной состав',
            'requirements': ['• Возраст: 18+', '• K/D: 8+', '• Коллекция: 50+', '• Аккаунт: 60+', '• SMS: 150 в неделю',
                             '• Energy: 2500 в неделю', '• Смена ника: 3 дня']},
        2: {'emoji': '🟡', 'name': 'NA KAIFE — академия',
            'requirements': ['• Возраст: 16+', '• K/D: M 6 на 100, W 5 на 100', '• Аккаунт: 50+', '• SMS: 300 в неделю',
                             '• Energy: 1500 в неделю', '• Смена ника: 7 дней']},
        3: {'emoji': '🟢', 'name': 'KAIF METRO',
            'requirements': ['• Возраст: 16+', '• K/D: 1.5+', '• Вынос: 1.5м', '• SMS: 300 в неделю',
                             '• Energy: 2000 в неделю', '• Смена ника: 7 дней']},
        4: {'emoji': '🟣', 'name': 'KAIF ESPORTS — турнирный состав',
            'requirements': ['• Возраст: 16+', '• SMS: 150 в неделю', '• Energy: 2000 в неделю', '• Смена ника: 3 дня',
                             '• Ответственность, дисциплина', '• Опыт турниров и праков']}
    }

    for clan in clans:
        clan_id, name, leader_id, leader_username, leader_name, deputy_id, deputy_username, deputy_name, _ = clan
        info = clan_data.get(clan_id, {})
        text += f'{info.get("emoji", "🔵")} {info.get("name", name)}\n'
        text += f'   👑 Лидер: {leader_name if leader_name else "❌ не назначен"}\n'
        text += f'   👤 Зам: {deputy_name if deputy_name else "❌ не назначен"}\n'
        text += f'   📋 Требования:\n'
        for req in info.get('requirements', []):
            text += f'   {req}\n'
        text += '\n─────────────────────\n\n'

    await callback.message.edit_text(text, reply_markup=back_button('back_to_main'))


# ============================================================
# 📞 КОНТАКТЫ
# ============================================================

@router.callback_query(F.data == 'contacts')
async def contacts(callback: CallbackQuery):
    await callback.answer()
    text = '📞 КОНТАКТЫ:\n\n👨‍💼 Менеджеры:\n   Xoma (@Xoma9991)\n   Катя (@Vibnot)'
    await callback.message.edit_text(text, reply_markup=back_button('back_to_main'))


# ============================================================
# 📋 ЗАЯВКИ В МОЙ КЛАН (ДЛЯ ЛИДЕРОВ/ЗАМОВ)
# ============================================================

@router.callback_query(F.data == 'my_clan_applications')
async def my_clan_applications(callback: CallbackQuery):
    await callback.answer()
    clan = await get_clan_by_user(callback.from_user.id)
    if not clan:
        return

    clan_id, name = clan[0], clan[1]
    apps = await get_clan_applications(clan_id)

    if not apps:
        await callback.message.edit_text(f'📋 ЗАЯВКИ В КЛАН {name}:\n\nПока нет ни одной заявки.',
                                         reply_markup=back_button('back_to_main'))
        return

    status_emoji = {'pending': '⏳ На рассмотрении', 'accepted': '✅ Принято', 'rejected': '❌ Отклонено',
                    'revoked': '⚠️ Отозвано'}
    text = f'📋 ЗАЯВКИ В КЛАН {name}:\n\n📌 Сверху — новые, снизу — старые.\n\n'

    buttons = []

    for idx, app in enumerate(apps, 1):
        app_id, user_id, username, clan_id_db, answers_json, photo_old, photo_new, has_photos, chat_id, status, created_at, reviewed_by, reviewed_at, clan_name = app
        emoji = ['1️⃣', '2️⃣', '3️⃣', '4️⃣', '5️⃣'][idx - 1] if idx <= 5 else '🔹'
        text += f'{emoji} #{app_id} — @{username} — {status_emoji.get(status, status)}\n'
        text += f'   📅 {created_at[:10]}, {created_at[11:16]}\n'
        text += f'   📸 {has_photos} фото\n'
        text += '\n'

    buttons.append([InlineKeyboardButton(text='🔙 Назад', callback_data='back_to_main')])
    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)

    await callback.message.edit_text(text, reply_markup=keyboard)


# ============================================================
# 📝 ПОДАТЬ АНКЕТУ
# ============================================================

@router.callback_query(F.data == 'apply')
async def apply_start(callback: CallbackQuery):
    await callback.answer()
    if await is_in_blacklist(callback.from_user.id):
        await callback.message.edit_text('🚫 Вы в чёрном списке кланов KAIF.\nОбратитесь к лидерам для разблокировки.',
                                         reply_markup=back_button('back_to_main'))
        return
    await callback.message.edit_text('Выберите клан для подачи заявки:', reply_markup=clan_choice())


# ============================================================
# 🎯 ВЫБОР КЛАНА
# ============================================================

@router.callback_query(F.data.startswith('clan_'))
async def select_clan(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    clan_id = int(callback.data.split('_')[1])
    clan = await get_clan(clan_id)
    if not clan:
        await callback.message.answer('❌ Клан не найден')
        return

    clan_id, name, leader_id, leader_username, leader_name, deputy_id, deputy_username, deputy_name, _ = clan

    if not leader_id and not deputy_id:
        await callback.message.edit_text('❌ В этом клане пока нет ответственных.\nЗаявки временно не принимаются.',
                                         reply_markup=back_button('back_to_main'))
        return

    existing = await get_pending_application(callback.from_user.id, clan_id)
    if existing:
        await callback.message.edit_text('⏳ У вас уже есть заявка в этот клан!\nДождитесь решения.',
                                         reply_markup=after_apply_buttons())
        return

    await state.update_data(clan_id=clan_id, clan_name=name)
    await state.set_state(ApplicationForm.waiting_for_answers)

    responsible = deputy_name if deputy_id else leader_name
    hint = f'ℹ️ Заявки принимает {responsible}\n\n'

    template = (
        'Имя: \n'
        'Возраст: \n'
        'Ник: \n'
        'ID: \n'
        'Часовой пояс (МСК): '
    )

    if name == "KAIF METRO":
        photo_text = (
            '📸 Теперь отправьте 2 фото:\n'
            '1️⃣ Скрин за ТЕКУЩИЙ сезон\n'
            '2️⃣ Скрин за ПРОШЛЫЙ сезон (если есть)'
        )
    else:
        photo_text = (
            '📸 Теперь отправьте 2 фото:\n'
            '1️⃣ Скрин за ПРОШЛЫЙ сезон\n'
            '2️⃣ Скрин за ТЕКУЩИЙ сезон'
        )

    await callback.message.edit_text(
        f'{hint}'
        '📋 СКОПИРУЙТЕ ШАБЛОН И ЗАПОЛНИТЕ:\n'
        '━━━━━━━━━━━━━━━━━━━━━━\n'
        f'{template}'
        '━━━━━━━━━━━━━━━━━━━━━━\n\n'
        '⚠️ Заполните все поля и отправьте одним сообщением.\n'
        'Каждое поле — с новой строки.\n\n'
        '📌 ПРИМЕР:\n'
        'Имя: Александр\n'
        'Возраст: 19\n'
        'Ник: KAIF_Pro\n'
        'ID: 123456789\n'
        'Часовой пояс (МСК): +0\n\n'
        f'{photo_text}',
        reply_markup=copy_template_button(template)
    )


# ============================================================
# 📥 ПОЛУЧЕНИЕ АНКЕТЫ
# ============================================================

@router.message(ApplicationForm.waiting_for_answers)
async def receive_application(message: Message, state: FSMContext):
    if await is_in_blacklist(message.from_user.id):
        await message.answer('🚫 Вы в чёрном списке.', reply_markup=back_button('back_to_main'))
        await state.clear()
        return

    data = await state.get_data()
    clan_id = data.get('clan_id')
    clan_name = data.get('clan_name')
    is_test = data.get('is_test_mode', False)

    text = message.text.strip()
    lines = [line.strip() for line in text.split('\n') if line.strip()]

    cleaned_lines = []
    for line in lines:
        if ':' in line or '：' in line:
            parts = re.split(r'[:：]', line, 1)
            if len(parts) == 2:
                cleaned_lines.append(parts[1].strip())
            else:
                cleaned_lines.append(line.strip())
        else:
            cleaned_lines.append(line.strip())

    if len(cleaned_lines) != 5:
        await message.answer(
            '❌ Нужно ровно 5 полей!\n\n'
            '📌 ПРАВИЛЬНЫЙ ФОРМАТ:\n'
            'Имя: Александр\n'
            'Возраст: 19\n'
            'Ник: KAIF_Pro\n'
            'ID: 123456789\n'
            'Часовой пояс (МСК): +0\n\n'
            '⚠️ Или просто 5 строк без названий полей:\n'
            'Александр\n'
            '19\n'
            'KAIF_Pro\n'
            '123456789\n'
            '+0',
            reply_markup=back_button('back_to_main')
        )
        return

    answers = {
        'name': cleaned_lines[0],
        'age': cleaned_lines[1],
        'nickname': cleaned_lines[2],
        'id': cleaned_lines[3],
        'timezone': cleaned_lines[4]
    }

    try:
        age = int(answers['age'])
        if age < 10 or age > 99:
            raise ValueError
    except:
        await message.answer(
            '❌ Возраст должен быть числом от 10 до 99!\n'
            'Попробуйте снова:',
            reply_markup=back_button('back_to_main')
        )
        return

    try:
        tz = int(answers['timezone'])
        if tz < -12 or tz > 12:
            raise ValueError
    except:
        await message.answer(
            '❌ Часовой пояс должен быть числом от -12 до +12!\n'
            'Попробуйте снова:',
            reply_markup=back_button('back_to_main')
        )
        return

    # Сохраняем в state, что это тестовый режим (если есть)
    await state.update_data(is_test_mode=is_test)

    # Добавляем заявку в БД
    if is_test:
        app_id = await add_application(message.from_user.id, 'test_user', clan_id, answers)
    else:
        app_id = await add_application(message.from_user.id, message.from_user.username or 'unknown', clan_id, answers)

    await state.update_data(app_id=app_id, answers=answers, is_test_mode=is_test)
    await state.set_state(ApplicationForm.waiting_photo_old)

    if clan_name == "KAIF METRO":
        await message.answer(
            f'✅ Анкета сохранена!\n\n'
            f'📋 ПРОВЕРЬТЕ ДАННЫЕ:\n'
            f'1. Имя: {answers["name"]}\n'
            f'2. Возраст: {answers["age"]}\n'
            f'3. Ник: {answers["nickname"]}\n'
            f'4. ID: {answers["id"]}\n'
            f'5. Часовой пояс (МСК): {answers["timezone"]}\n\n'
            f'📸 Теперь отправьте 2 фото:\n'
            f'1️⃣ Скрин за ТЕКУЩИЙ сезон\n'
            f'2️⃣ Скрин за ПРОШЛЫЙ сезон (если есть)\n\n'
            f'Второе фото можно пропустить.',
            reply_markup=photo_old_button()
        )
    else:
        await message.answer(
            f'✅ Анкета сохранена!\n\n'
            f'📋 ПРОВЕРЬТЕ ДАННЫЕ:\n'
            f'1. Имя: {answers["name"]}\n'
            f'2. Возраст: {answers["age"]}\n'
            f'3. Ник: {answers["nickname"]}\n'
            f'4. ID: {answers["id"]}\n'
            f'5. Часовой пояс (МСК): {answers["timezone"]}\n\n'
            f'📸 Теперь отправьте 2 фото:\n'
            f'1️⃣ Скрин за ПРОШЛЫЙ сезон\n'
            f'2️⃣ Скрин за ТЕКУЩИЙ сезон\n\n'
            f'Второе фото можно пропустить.',
            reply_markup=photo_old_button()
        )


# ============================================================
# 📸 ОТПРАВКА ФОТО 1
# ============================================================

@router.callback_query(F.data == 'send_photo_old')
async def send_photo_old(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    await state.set_state(ApplicationForm.waiting_photo_old)
    await callback.message.edit_text('📸 Отправьте фото 1.\nПросто пришлите фото в этот чат.')


# ============================================================
# 📥 ПОЛУЧЕНИЕ ФОТО 1
# ============================================================

@router.message(ApplicationForm.waiting_photo_old, F.photo)
async def receive_photo_old(message: Message, state: FSMContext):
    data = await state.get_data()
    app_id = data.get('app_id')
    is_test = data.get('is_test_mode', False)
    
    if not app_id:
        await message.answer('Ошибка. Попробуйте начать заново через /start')
        await state.clear()
        return

    await update_application_photo_old(app_id, message.photo[-1].file_id)
    await update_application_has_photos(app_id, 1)
    await state.update_data(photo_old=message.photo[-1].file_id, is_test_mode=is_test)
    await state.set_state(ApplicationForm.waiting_photo_new)

    clan_name = data.get('clan_name')
    if clan_name == "KAIF METRO":
        await message.answer(
            '✅ Фото 1 (текущий сезон) получено!\n'
            '📸 Теперь отправьте фото 2 (прошлый сезон, если есть)\n\n'
            'Или нажмите "Пропустить".',
            reply_markup=photo_new_button_with_skip()
        )
    else:
        await message.answer(
            '✅ Фото 1 (прошлый сезон) получено!\n'
            '📸 Теперь отправьте фото 2 (текущий сезон)\n\n'
            'Или нажмите "Пропустить".',
            reply_markup=photo_new_button_with_skip()
        )


# ============================================================
# 📸 ОТПРАВКА ФОТО 2
# ============================================================

@router.callback_query(F.data == 'send_photo_new')
async def send_photo_new(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    await state.set_state(ApplicationForm.waiting_photo_new)
    await callback.message.edit_text('📸 Отправьте фото 2.\nПросто пришлите фото в этот чат.')


# ============================================================
# 📥 ПОЛУЧЕНИЕ ФОТО 2 + УВЕДОМЛЕНИЕ ЛИДЕРУ И ЗАМУ
# ============================================================

@router.message(ApplicationForm.waiting_photo_new, F.photo)
async def receive_photo_new(message: Message, state: FSMContext):
    data = await state.get_data()
    app_id = data.get('app_id')
    clan_name = data.get('clan_name')
    photo_old = data.get('photo_old')
    answers = data.get('answers', {})
    is_test = data.get('is_test_mode', False)

    if not app_id:
        await message.answer('Ошибка. Попробуйте начать заново через /start')
        await state.clear()
        return

    photo_new = message.photo[-1].file_id
    await update_application_photo_new(app_id, photo_new)
    await update_application_has_photos(app_id, 2)

    await state.clear()

    try:
        clan = await get_clan_by_name(clan_name)
        if not clan:
            await message.answer('❌ Клан не найден.')
            return

        clan_id, name, leader_id, leader_username, leader_name, deputy_id, deputy_username, deputy_name, _ = clan

        text = (
            f'🔔 НОВАЯ ЗАЯВКА #{app_id} В КЛАН {clan_name}\n\n'
            f'От: @{message.from_user.username or "unknown"} (ID: {message.from_user.id})\n'
            f'Дата: {datetime.now().strftime("%d.%m.%Y, %H:%M")}\n\n'
            f'📝 АНКЕТА:\n'
            f'1. Имя: {answers.get("name", "")}\n'
            f'2. Возраст: {answers.get("age", "")}\n'
            f'3. Ник: {answers.get("nickname", "")}\n'
            f'4. ID: {answers.get("id", "")}\n'
            f'5. Часовой пояс (МСК): {answers.get("timezone", "")}\n\n'
            f'📸 Скринов: 2'
        )

        # Отправка лидеру
        if leader_id:
            try:
                await message.bot.send_media_group(
                    leader_id,
                    media=[
                        InputMediaPhoto(media=photo_old, caption=text),
                        InputMediaPhoto(media=photo_new)
                    ]
                )
                await message.bot.send_message(
                    leader_id,
                    "📌 Действия с заявкой:",
                    reply_markup=review_buttons(app_id)
                )
            except Exception as e:
                print(f"Ошибка отправки лидеру: {e}")

        # Отправка заму
        if deputy_id:
            try:
                await message.bot.send_media_group(
                    deputy_id,
                    media=[
                        InputMediaPhoto(media=photo_old, caption=text),
                        InputMediaPhoto(media=photo_new)
                    ]
                )
                await message.bot.send_message(
                    deputy_id,
                    "📌 Действия с заявкой:",
                    reply_markup=review_buttons(app_id)
                )
            except Exception as e:
                print(f"Ошибка отправки заму: {e}")

        # Если это тестовый режим
        if is_test:
            await message.answer(
                f'🧪 ТЕСТОВАЯ заявка #{app_id} отправлена на рассмотрение!\n'
                f'Ожидайте решения лидера или зама.',
                reply_markup=exit_test_button()
            )
        else:
            await message.answer(
                f'✅ Заявка #{app_id} отправлена на рассмотрение!\n'
                f'Ожидайте решения лидера или зама.',
                reply_markup=after_apply_buttons()
            )

    except Exception as e:
        await message.answer(f'❌ Ошибка: {e}\nЗаявка сохранена.')


# ============================================================
# 💬 СВЯЗАТЬСЯ С КАНДИДАТОМ (С СОХРАНЕНИЕМ ССЫЛОК)
# ============================================================

@router.callback_query(F.data.startswith('contact_'))
async def contact_application(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    app_id = int(callback.data.split('_')[1])

    clan = await get_clan_by_user(callback.from_user.id)
    if not clan:
        await callback.message.answer('⛔ У вас нет прав на это действие')
        return

    app = await get_application_by_id(app_id)
    if not app:
        await callback.message.answer('❌ Заявка не найдена')
        return

    await state.update_data(contact_app_id=app_id)

    clan_id, clan_name = clan[0], clan[1]

    link = await get_clan_link(clan_id)

    if link:
        await callback.message.answer(
            f'📩 Сообщение кандидату @{app[2]}\n\n'
            f'Текущая ссылка на чат: {link}\n\n'
            f'Что хотите сделать?',
            reply_markup=contact_with_link(app_id, link)
        )
    else:
        await callback.message.answer(
            f'📩 Сообщение кандидату @{app[2]}\n\n'
            f'У клана ещё нет ссылки на чат.\n'
            f'Сначала добавьте её.',
            reply_markup=contact_menu(app_id)
        )


# ============================================================
# 🔗 ДОБАВИТЬ ССЫЛКУ
# ============================================================

@router.callback_query(F.data.startswith('add_link_'))
async def add_link(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    app_id = int(callback.data.split('_')[2])

    await state.update_data(contact_app_id=app_id)
    await state.set_state(ApplicationForm.waiting_contact_message)
    await state.update_data(link_action='add')

    await callback.message.edit_text(
        '✏️ Введите ссылку на чат клана:\n'
        'Например: https://t.me/joinchat/xxxxx\n\n'
        'Или отправьте @username чата.',
        reply_markup=back_button('back_to_main')
    )


# ============================================================
# ✏️ ИЗМЕНИТЬ ССЫЛКУ
# ============================================================

@router.callback_query(F.data.startswith('edit_link_'))
async def edit_link(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    app_id = int(callback.data.split('_')[2])

    await state.update_data(contact_app_id=app_id)
    await state.set_state(ApplicationForm.waiting_contact_message)
    await state.update_data(link_action='edit')

    await callback.message.edit_text(
        '✏️ Введите новую ссылку на чат клана:\n'
        'Например: https://t.me/joinchat/xxxxx\n\n'
        'Или отправьте @username чата.',
        reply_markup=back_button('back_to_main')
    )


# ============================================================
# 📤 ОТПРАВИТЬ СООБЩЕНИЕ (С ФОРМАТИРОВАННЫМ ТЕКСТОМ)
# ============================================================

@router.callback_query(F.data.startswith('send_message_'))
async def send_message(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    app_id = int(callback.data.split('_')[2])

    clan = await get_clan_by_user(callback.from_user.id)
    if not clan:
        await callback.message.answer('⛔ У вас нет прав на это действие')
        return

    app = await get_application_by_id(app_id)
    if not app:
        await callback.message.answer('❌ Заявка не найдена')
        return

    clan_id, clan_name, leader_id, leader_username, leader_name, deputy_id, deputy_username, deputy_name, _ = clan
    link = await get_clan_link(clan_id)

    if not link:
        await callback.message.answer('❌ Ссылка на чат не найдена. Сначала добавьте её.')
        return

    # ID ЛИДЕРОВ КЛАНОВ (ДЛЯ ИГРЫ)
    clan_ids = {
        'KAIF': '51656781871',
        'KAIF ESPORTS': '51600572333',
        'KAIF METRO': '51954255028',
        'NA KAIFE': '51768659282'
    }
    clan_id_game = clan_ids.get(clan_name, 'не указан')

    # Определяем, кто отправляет
    if callback.from_user.id == leader_id:
        sender = f"👑 Лидер клана {clan_name} — {leader_name}"
    elif callback.from_user.id == deputy_id:
        sender = f"👤 Зам клана {clan_name} — {deputy_name}"
    else:
        sender = f"Администрация {clan_name}"

    message_text = (
        f'🎉 ПОЗДРАВЛЯЕМ!\n\n'
        f'Вы прошли отбор и официально приняты в клан {clan_name}!\n\n'
        f'Добро пожаловать в нашу дружную команду! Мы рады, что ты с нами. Впереди — совместные игры, турниры, тренировки и новые достижения.\n\n'
        f'🔥 Сделай ник с припиской KAIF\n\n'
        f'📌 Ссылка на чат клана: {link}\n\n'
        f'🆔 ID лидера для подачи заявки в игре: {clan_id_game}\n\n'
        f'📩 Отправил: {sender}\n\n'
        f'С уважением, администрация {clan_name} ❤️'
    )

    try:
        await callback.bot.send_message(
            app[1],
            message_text
        )
        await callback.message.answer(f'✅ Сообщение отправлено кандидату @{app[2]}!')
    except Exception as e:
        await callback.message.answer(f'❌ Не удалось отправить сообщение. Ошибка: {e}')

    await state.clear()


# ============================================================
# 📨 ПОЛУЧЕНИЕ СООБЩЕНИЯ (ДЛЯ ДОБАВЛЕНИЯ ССЫЛКИ)
# ============================================================

@router.message(ApplicationForm.waiting_contact_message)
async def handle_contact_message(message: Message, state: FSMContext):
    data = await state.get_data()
    app_id = data.get('contact_app_id')
    link_action = data.get('link_action')

    if not app_id:
        await message.answer('❌ Ошибка. Попробуйте снова.', reply_markup=main_menu())
        await state.clear()
        return

    app = await get_application_by_id(app_id)
    if not app:
        await message.answer('❌ Заявка не найдена')
        await state.clear()
        return

    # Если это действие по добавлению/изменению ссылки
    if link_action in ['add', 'edit']:
        # Получаем клан из заявки
        clan_id = app[3]  # clan_id находится на позиции 3 в кортеже заявки
        link = message.text.strip()

        print(f"🔍 СОХРАНЯЕМ ССЫЛКУ: clan_id={clan_id}, link={link}")

        await set_clan_link(clan_id, link)

        action_text = 'добавлена' if link_action == 'add' else 'обновлена'
        await message.answer(f'✅ Ссылка на чат клана {action_text}!\n\nТекущая ссылка: {link}')
        await state.clear()
        return

    # Если это просто сообщение для кандидата
    try:
        await message.bot.send_message(
            app[1],
            f'📩 Сообщение от лидера клана {app[12]}:\n\n{message.text}'
        )
        await message.answer(f'✅ Сообщение отправлено кандидату @{app[2]}!')
    except Exception as e:
        await message.answer(f'❌ Не удалось отправить сообщение. Ошибка: {e}')

    await state.clear()


# ============================================================
# 📊 МОИ ЗАЯВКИ
# ============================================================

@router.callback_query(F.data == 'my_applications')
async def my_applications(callback: CallbackQuery):
    await callback.answer()
    apps = await get_user_applications(callback.from_user.id)
    if not apps:
        await callback.message.edit_text(
            '📊 У вас пока нет заявок.\nПодайте анкету через 📝 Подать анкету',
            reply_markup=back_button('back_to_main')
        )
        return

    status_emoji = {
        'pending': '⏳ На рассмотрении',
        'accepted': '✅ Принято',
        'rejected': '❌ Отклонено',
        'revoked': '⚠️ Отозвано'
    }

    text = '📊 ВАШИ ЗАЯВКИ:\n\n'
    buttons = []

    for app in apps:
        app_id, user_id, username, clan_id, answers_json, photo_old, photo_new, has_photos, chat_id, status, created_at, reviewed_by, reviewed_at, clan_name = app

        text += f'{status_emoji.get(status, status)} в клан {clan_name}\n'
        text += f'   От: {created_at[:10]}\n'
        text += f'   📸 {has_photos} фото\n'

        if status == 'pending':
            buttons.append(
                [InlineKeyboardButton(text=f'❌ Отозвать заявку #{app_id}', callback_data=f'revoke_{app_id}')])
        text += '\n'

    text += '💡 Если заявка не рассмотрена, вы можете её отозвать.'
    buttons.append([InlineKeyboardButton(text='🔙 Назад', callback_data='back_to_main')])
    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)

    await callback.message.edit_text(text, reply_markup=keyboard)


# ============================================================
# ❌ ОТОЗВАТЬ ЗАЯВКУ
# ============================================================

@router.callback_query(F.data.startswith('revoke_'))
async def revoke_application_handler(callback: CallbackQuery):
    await callback.answer()
    app_id = int(callback.data.split('_')[1])
    await revoke_application(app_id)
    await callback.message.edit_text('⚠️ Заявка отозвана.', reply_markup=after_apply_buttons())


# ============================================================
# ✅ ПРИНЯТЬ ЗАЯВКУ
# ============================================================

@router.callback_query(F.data.startswith('accept_'))
async def accept_application(callback: CallbackQuery):
    await callback.answer()
    app_id = int(callback.data.split('_')[1])
    clan = await get_clan_by_user(callback.from_user.id)
    if not clan:
        await callback.message.answer('⛔ У вас нет прав на это действие')
        return

    await update_application_status(app_id, 'accepted', callback.from_user.id)
    await callback.message.edit_reply_markup(reply_markup=None)

    await callback.message.answer(
        f'🎉 Заявка #{app_id} ПРИНЯТА!\n\n'
        f'Теперь вы можете отправить кандидату ссылку на вступление.',
        reply_markup=contact_button(app_id)
    )

    app = await get_application_by_id(app_id)
    if app:
        try:
            await callback.bot.send_message(
                app[1],
                f'🎉 ПОЗДРАВЛЯЕМ!\nВаша заявка в клан {app[12]} ПРИНЯТА!\n'
                f'Ожидайте сообщение от лидера/зама.'
            )
        except:
            pass


# ============================================================
# ❌ ОТКЛОНИТЬ ЗАЯВКУ
# ============================================================

@router.callback_query(F.data.startswith('reject_'))
async def reject_application(callback: CallbackQuery):
    await callback.answer()
    app_id = int(callback.data.split('_')[1])
    clan = await get_clan_by_user(callback.from_user.id)
    if not clan:
        await callback.message.answer('⛔ У вас нет прав на это действие')
        return

    await update_application_status(app_id, 'rejected', callback.from_user.id)
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer('❌ Заявка ОТКЛОНЕНА.')

    app = await get_application_by_id(app_id)
    if app:
        try:
            await callback.bot.send_message(
                app[1],
                f'😔 К сожалению, ваша заявка в клан {app[12]} отклонена.\nВы можете подать заявку в другой клан или попробовать позже.'
            )
        except:
            pass


# ============================================================
# 👥 УПРАВЛЕНИЕ РУКОВОДИТЕЛЯМИ
# ============================================================

@router.callback_query(F.data == 'admin_manage_roles')
async def admin_manage_roles(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    await callback.message.edit_text(
        '👥 Управление руководителями\n\n'
        'Здесь вы можете назначить или удалить лидера/зама для любого клана.',
        reply_markup=manage_roles_menu()
    )


# ============================================================
# ➕ НАЗНАЧИТЬ ЛИДЕРА
# ============================================================

@router.callback_query(F.data == 'role_assign_leader')
async def role_assign_leader(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    await state.update_data(role_type='leader')
    await callback.message.edit_text(
        '👥 Назначение лидера\n\n'
        'Выберите действие:',
        reply_markup=assign_choice_menu()
    )


# ============================================================
# ➕ НАЗНАЧИТЬ ЗАМА
# ============================================================

@router.callback_query(F.data == 'role_assign_deputy')
async def role_assign_deputy(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    await state.update_data(role_type='deputy')
    await callback.message.edit_text(
        '👥 Назначение зама\n\n'
        'Выберите действие:',
        reply_markup=assign_choice_menu()
    )


# ============================================================
# 📋 ВЫБРАТЬ ИЗ СУЩЕСТВУЮЩИХ
# ============================================================

@router.callback_query(F.data == 'assign_from_existing')
async def assign_from_existing(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    data = await state.get_data()
    role_type = data.get('role_type', 'leader')
    role_name = 'лидером' if role_type == 'leader' else 'замом'

    clans = await get_clans()
    leaders = []
    for clan in clans:
        clan_id, name, leader_id, leader_username, leader_name, deputy_id, deputy_username, deputy_name, _ = clan
        if leader_id:
            leaders.append({
                'id': leader_id,
                'username': leader_username,
                'name': leader_name,
                'clan': name,
                'clan_id': clan_id,
                'role': 'Лидер'
            })
        if deputy_id:
            leaders.append({
                'id': deputy_id,
                'username': deputy_username,
                'name': deputy_name,
                'clan': name,
                'clan_id': clan_id,
                'role': 'Зам'
            })

    if not leaders:
        await callback.message.edit_text(
            '❌ Нет существующих руководителей.\n'
            'Используйте "Ввести нового пользователя".',
            reply_markup=assign_choice_menu()
        )
        return

    text = f'👥 Выберите руководителя для назначения {role_name}:\n\n'
    for idx, leader in enumerate(leaders[:10], 1):
        emoji = '👑' if leader['role'] == 'Лидер' else '👤'
        text += f"{idx}. {leader['name']} (@{leader['username']}) — {emoji} {leader['role']} {leader['clan']}\n"
    if len(leaders) > 10:
        text += f"\n... и ещё {len(leaders) - 10} человек"

    await callback.message.edit_text(text, reply_markup=select_existing_leader_buttons(leaders, role_type))


# ============================================================
# ✏️ ВВЕСТИ НОВОГО ПОЛЬЗОВАТЕЛЯ
# ============================================================

@router.callback_query(F.data == 'assign_from_new')
async def assign_from_new(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    await state.set_state(RoleForm.waiting_user_id)
    await callback.message.edit_text(
        '✏️ Введите данные нового пользователя:\n\n'
        '1️⃣ Telegram ID (число):\n'
        'Пример: 123456789\n\n'
        '[❌ Отмена]',
        reply_markup=cancel_button()
    )


# ============================================================
# 👤 ВЫБОР СУЩЕСТВУЮЩЕГО РУКОВОДИТЕЛЯ
# ============================================================

@router.callback_query(F.data.startswith('select_existing_'))
async def select_existing_leader(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    parts = callback.data.split('_')
    user_id = int(parts[2])
    clan_id = int(parts[3])

    data = await state.get_data()
    role_type = data.get('role_type', 'leader')
    role_name = 'лидером' if role_type == 'leader' else 'замом'

    clans = await get_clans()
    user_info = None
    for clan in clans:
        clan_id_db, name, leader_id, leader_username, leader_name, deputy_id, deputy_username, deputy_name, _ = clan
        if leader_id == user_id:
            user_info = {'id': leader_id, 'username': leader_username, 'name': leader_name}
            break
        if deputy_id == user_id:
            user_info = {'id': deputy_id, 'username': deputy_username, 'name': deputy_name}
            break

    if not user_info:
        await callback.message.answer('❌ Пользователь не найден')
        return

    await state.update_data(selected_user_id=user_id, selected_username=user_info['username'],
                            selected_name=user_info['name'])

    await callback.message.edit_text(
        f'👤 Выбран: {user_info["name"]} (@{user_info["username"]})\n\n'
        f'Выберите клан для назначения {role_name}:',
        reply_markup=select_clan_for_role_buttons(clans, role_type, user_id, user_info['username'], user_info['name'])
    )


# ============================================================
# 📌 НАЗНАЧИТЬ В КЛАН
# ============================================================

@router.callback_query(F.data.startswith('assign_to_clan_'))
async def assign_to_clan(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    parts = callback.data.split('_')
    role_type = parts[3]
    clan_id = int(parts[4])
    user_id = int(parts[5])
    username = parts[6]
    name = parts[7]

    clan = await get_clan(clan_id)
    if not clan:
        await callback.message.answer('❌ Клан не найден')
        return

    clans = await get_clans()
    for c in clans:
        c_id, c_name, leader_id, leader_username, leader_name, deputy_id, deputy_username, deputy_name, _ = c
        if leader_id == user_id:
            await remove_clan_leader(c_id)
        if deputy_id == user_id:
            await remove_clan_deputy(c_id)

    if role_type == 'leader':
        await update_clan_leader(clan_id, user_id, username, name)
        await callback.message.edit_text(
            f'✅ {name} (@{username}) назначен лидером клана {clan[1]}!\n'
            f'Старая должность автоматически удалена.'
        )
    else:
        await update_clan_deputy(clan_id, user_id, username, name)
        await callback.message.edit_text(
            f'✅ {name} (@{username}) назначен замом клана {clan[1]}!\n'
            f'Старая должность автоматически удалена.'
        )

    await state.clear()
    await callback.message.answer('👥 Управление руководителями\n\nВыберите действие:', reply_markup=manage_roles_menu())


# ============================================================
# 🗑 УДАЛИТЬ ЛИДЕРА
# ============================================================

@router.callback_query(F.data == 'role_remove_leader')
async def role_remove_leader(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    await callback.message.edit_text('Выберите клан для удаления лидера:', reply_markup=clan_choice_for_roles())


# ============================================================
# 🗑 УДАЛИТЬ ЗАМА
# ============================================================

@router.callback_query(F.data == 'role_remove_deputy')
async def role_remove_deputy(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    await callback.message.edit_text('Выберите клан для удаления зама:', reply_markup=clan_choice_for_roles())


# ============================================================
# 🔄 ВЫБОР КЛАНА ДЛЯ УДАЛЕНИЯ
# ============================================================

@router.callback_query(F.data.startswith('role_clan_'))
async def role_remove_select_clan(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    clan_id = int(callback.data.split('_')[2])
    clan = await get_clan(clan_id)
    if not clan:
        await callback.message.answer('❌ Клан не найден')
        return

    await callback.message.edit_text(
        f'⚠️ Вы уверены, что хотите удалить руководителя из клана {clan[1]}?',
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text='✅ Да, удалить', callback_data=f'role_confirm_remove_{clan_id}')],
            [InlineKeyboardButton(text='🔙 Назад', callback_data='back_to_roles')],
        ])
    )


# ============================================================
# ✅ ПОДТВЕРЖДЕНИЕ УДАЛЕНИЯ
# ============================================================

@router.callback_query(F.data.startswith('role_confirm_remove_'))
async def role_confirm_remove(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    clan_id = int(callback.data.split('_')[3])
    clan = await get_clan(clan_id)

    if clan[2]:
        await remove_clan_leader(clan_id)
        await callback.message.edit_text(f'✅ Лидер удалён из клана {clan[1]}')
    elif clan[5]:
        await remove_clan_deputy(clan_id)
        await callback.message.edit_text(f'✅ Зам удалён из клана {clan[1]}')
    else:
        await callback.message.edit_text(f'❌ В клане {clan[1]} нет руководителей для удаления.')

    await callback.message.answer('👥 Управление руководителями\n\nВыберите действие:', reply_markup=manage_roles_menu())


# ============================================================
# 📋 СПИСОК РУКОВОДИТЕЛЕЙ
# ============================================================

@router.callback_query(F.data == 'role_list')
async def role_list(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    clans = await get_clans()
    text = '👥 ТЕКУЩИЕ РУКОВОДИТЕЛИ:\n\n'
    emojis = {1: '🔴', 2: '🟡', 3: '🟢', 4: '🟣'}

    for clan in clans:
        clan_id, name, leader_id, leader_username, leader_name, deputy_id, deputy_username, deputy_name, _ = clan
        emoji = emojis.get(clan_id, '🔵')

        text += f'{emoji} {name}:\n'
        text += f'   👑 Лидер: {leader_name if leader_name else "❌ не назначен"}'
        if leader_username:
            text += f' (@{leader_username})'
        if leader_id:
            text += f' (ID: {leader_id})'
        text += '\n'

        text += f'   👤 Зам: {deputy_name if deputy_name else "❌ не назначен"}'
        if deputy_username:
            text += f' (@{deputy_username})'
        if deputy_id:
            text += f' (ID: {deputy_id})'
        text += '\n\n'

    await callback.message.edit_text(text, reply_markup=manage_roles_menu())


# ============================================================
# 📤 АДМИН: ЭКСПОРТ EXCEL (С ЦВЕТАМИ)
# ============================================================

@router.callback_query(F.data == 'admin_export')
async def admin_export(callback: CallbackQuery):
    print("🔍 НАЖАТА КНОПКА CSV!")
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        apps = await get_all_applications()
        if not apps:
            await callback.message.answer('❌ Нет заявок для экспорта')
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Заявки"

        headers = [
            'ID', 'User ID', 'Username', 'Клан', 'Имя', 'Возраст',
            'Ник', 'ID игровой', 'Часовой пояс', 'Скрин прошлый',
            'Скрин текущий', 'Статус', 'Дата создания',
            'Кто одобрил (ID)', 'Кто одобрил (Username)', 'Дата одобрения'
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        status_colors = {
            'accepted': PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
            'rejected': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            'pending': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
            'revoked': PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid"),
        }

        for app in apps:
            (app_id, user_id, username, clan_name, answers_json,
             photo_old, photo_new, has_photos, chat_id, status,
             created_at, reviewed_by, reviewed_at) = app

            answers = json.loads(answers_json)

            is_test = username == 'test_user' or 'Тест' in answers.get('name', '')

            status_ru = {
                'pending': '⏳ На рассмотрении',
                'accepted': '✅ Принято',
                'rejected': '❌ Отклонено',
                'revoked': '⚠️ Отозвано'
            }.get(status, status)

            if is_test:
                status_ru = '🧪 ' + status_ru + ' (ТЕСТ)'

            reviewer_username = ''
            if reviewed_by:
                try:
                    async with aiosqlite.connect(DB_PATH) as db:
                        async with db.execute(
                            'SELECT username FROM applications WHERE reviewed_by = ? LIMIT 1',
                            (reviewed_by,)
                        ) as cursor:
                            result = await cursor.fetchone()
                            if result:
                                reviewer_username = result[0]
                except:
                    reviewer_username = str(reviewed_by)

            row = [
                app_id, user_id, f'@{username}', clan_name,
                answers.get('name', ''), answers.get('age', ''),
                answers.get('nickname', ''), answers.get('id', ''),
                answers.get('timezone', ''),
                '✅' if photo_old else '❌',
                '✅' if photo_new else '❌',
                status_ru,
                created_at[:10] if created_at else '',
                reviewed_by if reviewed_by else '',
                reviewer_username,
                reviewed_at[:16] if reviewed_at else ''
            ]
            ws.append(row)

            row_num = ws.max_row

            if is_test:
                fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
                font_color = "FFFFFF"
            else:
                if status == 'accepted':
                    fill = status_colors['accepted']
                    font_color = "000000"
                elif status == 'rejected':
                    fill = status_colors['rejected']
                    font_color = "FFFFFF"
                elif status == 'pending':
                    fill = status_colors['pending']
                    font_color = "000000"
                else:
                    fill = status_colors['revoked']
                    font_color = "000000"

            status_cell = ws.cell(row=row_num, column=12)
            status_cell.fill = fill
            status_cell.font = Font(bold=True, color=font_color, size=10)
            status_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = max_length + 3

        ws.freeze_panes = 'A2'

        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from aiogram.types import BufferedInputFile
        await callback.message.answer_document(
            document=BufferedInputFile(output.getvalue(), filename='заявки.xlsx'),
            caption='📊 Все заявки в формате Excel с цветами!'
        )

    except Exception as e:
        await callback.message.answer(f'❌ Ошибка при экспорте: {e}')


# ============================================================
# 📊 АДМИН: СТАТИСТИКА
# ============================================================

@router.callback_query(F.data == 'admin_stats')
async def admin_stats(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    stats, by_clan = await get_statistics()
    total, pending, accepted, rejected, revoked = stats

    text = f'📊 СТАТИСТИКА ЗАЯВОК:\n\nВсего: {total}\n⏳ На рассмотрении: {pending}\n✅ Принято: {accepted}\n❌ Отклонено: {rejected}\n⚠️ Отозвано: {revoked}\n\nПо кланам:\n'
    for clan_name, count in by_clan:
        text += f'   {clan_name}: {count} заявок\n'

    await callback.message.edit_text(text, reply_markup=admin_menu())


# ============================================================
# 👥 АДМИН: ЧЁРНЫЙ СПИСОК
# ============================================================

@router.callback_query(F.data == 'admin_blacklist')
async def admin_blacklist(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    blacklist = await get_blacklist()
    if not blacklist:
        text = '👥 Чёрный список пуст.'
    else:
        text = '👥 ЧЁРНЫЙ СПИСОК:\n\n'
        for item in blacklist:
            text += f'ID: {item[1]}\nПричина: {item[2]}\nДобавлен: {item[4][:10] if item[4] else "неизвестно"}\n\n'

    await callback.message.edit_text(text, reply_markup=admin_menu())


# ============================================================
# 🗑 ОЧИСТИТЬ ТЕСТОВЫЕ ЗАЯВКИ (ДЛЯ АДМИНОВ)
# ============================================================

@router.callback_query(F.data == 'admin_clear_test')
async def admin_clear_test(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute("SELECT COUNT(*) FROM applications WHERE username = 'test_user'") as cursor:
            count = await cursor.fetchone()
            count = count[0] if count else 0

    if count == 0:
        await callback.message.edit_text(
            '🧪 Нет тестовых заявок для удаления.',
            reply_markup=admin_menu()
        )
        return

    await callback.message.edit_text(
        f'⚠️ ВЫ УВЕРЕНЫ?\n\n'
        f'Будут удалены ВСЕ тестовые заявки (с пометкой "ТЕСТ").\n'
        f'Найдено: {count} тестовых заявок.\n'
        f'Обычные заявки останутся нетронутыми.\n\n'
        f'Это действие нельзя отменить!',
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text='✅ Да, удалить все', callback_data='confirm_clear_test')],
            [InlineKeyboardButton(text='❌ Отмена', callback_data='back_to_admin')],
        ])
    )


@router.callback_query(F.data == 'confirm_clear_test')
async def confirm_clear_test(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    await clear_test_applications()

    await callback.message.edit_text(
        '✅ Все тестовые заявки удалены!\n\n'
        'Обычные заявки остались нетронутыми.',
        reply_markup=admin_menu()
    )
