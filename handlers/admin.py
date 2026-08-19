import json
import os
from datetime import datetime
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup
from aiogram.fsm.context import FSMContext

from config import ADMIN_IDS
from database import *
from keyboards import *
from .start import RoleForm, ClanManagementForm

router = Router()


# ============================================================
# 👤 ОБРАБОТКА ВВОДА ИМЕНИ
# ============================================================

@router.message(RoleForm.waiting_name)
async def process_new_user_name(message: Message, state: FSMContext):
    """Обработка ввода имени нового пользователя и выбор клана"""
    text = message.text.strip()
    if text.lower() in ['пропустить', 'skip']:
        name = None
    else:
        name = text
    
    data = await state.get_data()
    user_id = data.get('new_user_id')
    username = data.get('new_username')
    role_type = data.get('role_type', 'leader')
    role_name = 'лидером' if role_type == 'leader' else 'замом'
    
    await state.update_data(new_name=name)
    await state.update_data(pending_name=name)
    await state.update_data(new_username=username)  # ← СОХРАНЯЕМ USERNAME
    
    clans = await get_clans()
    text_msg = f'👤 Новый пользователь:\n'
    text_msg += f'ID: {user_id}\n'
    text_msg += f'Username: {username if username else "❌"}\n'
    text_msg += f'Имя: {name if name else "❌"}\n\n'
    text_msg += f'Выберите клан для назначения {role_name}:'
    
    await state.set_state(RoleForm.waiting_clan_id)
    
    await message.answer(
        text_msg,
        reply_markup=select_clan_for_role_buttons_simple(clans, role_type, user_id, username or '')
    )

# ============================================================
# 👥 УПРАВЛЕНИЕ РУКОВОДИТЕЛЯМИ
# ============================================================

@router.callback_query(F.data == 'admin_manage_roles')
async def admin_manage_roles(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    await callback.message.edit_text(
        '👥 Управление руководителями\n\n'
        'Здесь вы можете назначить или удалить лидера/зама для любого клана.',
        reply_markup=manage_roles_menu()
    )


@router.callback_query(F.data == 'role_assign_leader')
async def role_assign_leader(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    await state.update_data(role_type='leader')
    await callback.message.edit_text(
        '👥 Назначение лидера\n\n'
        'Выберите действие:',
        reply_markup=assign_choice_menu()
    )


@router.callback_query(F.data == 'role_assign_deputy')
async def role_assign_deputy(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    await state.update_data(role_type='deputy')
    await callback.message.edit_text(
        '👥 Назначение зама\n\n'
        'Выберите действие:',
        reply_markup=assign_choice_menu()
    )


@router.callback_query(F.data == 'assign_from_existing')
async def assign_from_existing(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    data = await state.get_data()
    role_type = data.get('role_type', 'leader')
    role_name = 'лидером' if role_type == 'leader' else 'замом'

    clans = await get_clans()
    leaders = []
    for clan in clans:
        if len(clan) < 9:
            continue
        
        # ПРАВИЛЬНЫЙ ПОРЯДОК ИЗ ЛОГОВ:
        # [0]=id, [1]=name, [2]=leader_id, [3]=leader_username,
        # [4]=leader_name, [5]=deputy_id, [6]=deputy_username,
        # [7]=deputy_name, [8]=created_at, [9]=is_active, [10]=emoji
        
        clan_id = clan[0]
        clan_name = clan[1]
        
        leader_id = clan[2] if len(clan) > 2 else None
        leader_username = clan[3] if len(clan) > 3 else ''
        leader_name = clan[4] if len(clan) > 4 else ''
        deputy_id = clan[5] if len(clan) > 5 else None
        deputy_username = clan[6] if len(clan) > 6 else ''
        deputy_name = clan[7] if len(clan) > 7 else ''
        
        # ✅ ФИКС: если leader_name пустой, используем leader_username как имя
        if not leader_name:
            leader_name = leader_username or '❌'
        if not deputy_name:
            deputy_name = deputy_username or '❌'
        
        # ✅ ФИКС: если leader_username содержит имя (кириллица), значит это имя, а не username
        if leader_username and any(ord(c) > 127 for c in leader_username):
            leader_name = leader_username
            leader_username = ''
        
        if deputy_username and any(ord(c) > 127 for c in deputy_username):
            deputy_name = deputy_username
            deputy_username = ''
        
        if leader_id and isinstance(leader_id, int):
            leaders.append({
                'id': leader_id,
                'username': leader_username or '',
                'name': leader_name or '❌',
                'clan': clan_name,
                'clan_id': clan_id,
                'role': 'Лидер'
            })
        if deputy_id and isinstance(deputy_id, int):
            leaders.append({
                'id': deputy_id,
                'username': deputy_username or '',
                'name': deputy_name or '❌',
                'clan': clan_name,
                'clan_id': clan_id,
                'role': 'Зам'
            })

    if not leaders:
        await callback.message.answer(
            '❌ Нет существующих руководителей.\n'
            'Используйте "Ввести нового пользователя".',
            reply_markup=assign_choice_menu()
        )
        return

    text = f'👥 Выберите руководителя для назначения {role_name}:\n\n'
    for idx, leader in enumerate(leaders[:10], 1):
        emoji = '👑' if leader['role'] == 'Лидер' else '👤'
        text += f"{idx}. {leader['name']} (@{leader['username']}) — {emoji} {leader['role']} {leader['clan']}\n"
    if len(leaders) > 10:
        text += f"\n... и ещё {len(leaders) - 10} человек"

    await callback.message.edit_text(text, reply_markup=select_existing_leader_buttons(leaders, role_type))

@router.callback_query(F.data == 'assign_from_new')
async def assign_from_new(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    await state.set_state(RoleForm.waiting_user_id)
    await callback.message.edit_text(
        '✏️ Введите данные нового пользователя:\n\n'
        '1️⃣ Telegram ID (число):\n'
        'Пример: 123456789\n\n'
        '[❌ Отмена]',
        reply_markup=cancel_button()
    )


@router.message(RoleForm.waiting_user_id)
async def process_new_user_id(message: Message, state: FSMContext):
    """Обработка ID нового пользователя"""
    try:
        user_id = int(message.text.strip())
    except ValueError:
        await message.answer('❌ Введите корректный ID (число):')
        return
    
    await state.update_data(new_user_id=user_id)
    await state.set_state(RoleForm.waiting_username)
    await message.answer(
        f'✅ ID: {user_id}\n\n'
        '2️⃣ Введите USERNAME (без @):\n'
        'Пример: username\n\n'
        'Или отправьте "пропустить"',
        reply_markup=cancel_button()
    )


@router.message(RoleForm.waiting_username)
async def process_new_username(message: Message, state: FSMContext):
    """Обработка username нового пользователя"""
    text = message.text.strip()
    if text.lower() in ['пропустить', 'skip']:
        username = None
    else:
        username = text.replace('@', '')
    
    await state.update_data(new_username=username)
    await state.set_state(RoleForm.waiting_name)
    await message.answer(
        f'✅ Username: {username if username else "❌"}\n\n'
        '3️⃣ Введите ИМЯ пользователя:\n'
        'Пример: Антон\n\n'
        'Или отправьте "пропустить"',
        reply_markup=cancel_button()
    )


@router.callback_query(F.data.startswith('select_existing_'))
async def select_existing_leader(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    parts = callback.data.split('_')
    
    try:
        user_id = int(parts[2])
    except (ValueError, IndexError):
        await callback.message.answer('❌ Ошибка: неверный формат ID')
        return
    
    try:
        clan_id = int(parts[3])
    except (ValueError, IndexError):
        await callback.message.answer('❌ Ошибка: неверный формат клана')
        return

    data = await state.get_data()
    role_type = data.get('role_type', 'leader')
    role_name = 'лидером' if role_type == 'leader' else 'замом'

    clans = await get_clans()
    user_info = None
    for clan in clans:
        if len(clan) < 9:
            continue
        # ПРАВИЛЬНЫЙ ПОРЯДОК:
        # [0]=id, [1]=name, [2]=leader_id, [3]=leader_username,
        # [4]=leader_name, [5]=deputy_id, [6]=deputy_username, [7]=deputy_name
        
        if clan[2] == user_id:  # leader_id на [2]
            user_info = {
                'id': clan[2],
                'username': clan[3] or '',
                'name': clan[4] or clan[3] or 'Пользователь'
            }
            break
        if clan[5] == user_id:  # deputy_id на [5]
            user_info = {
                'id': clan[5],
                'username': clan[6] or '',
                'name': clan[7] or clan[6] or 'Пользователь'
            }
            break

    if not user_info:
        await callback.message.answer('❌ Пользователь не найден')
        return

    await state.update_data(
        new_user_id=user_id,
        selected_username=user_info['username'],
        selected_name=user_info['name'],
        pending_name=user_info['name'],
        new_name=user_info['name'],
        new_username=user_info['username']
    )

    await callback.message.edit_text(
        f'👤 Выбран: {user_info["name"]} (@{user_info["username"]})\n\n'
        f'Выберите клан для назначения {role_name}:',
        reply_markup=select_clan_for_role_buttons(clans, role_type, user_id, user_info['username'], user_info['name'])
    )


@router.callback_query(F.data.startswith('assign_to_clan_'))
async def assign_to_clan(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    parts = callback.data.split('_')
    
    try:
        # Формат: assign_to_clan_{role_type}_{clan_id}_{user_id}
        role_type = parts[3]
        clan_id = int(parts[4])
        user_id = int(parts[5])
    except (ValueError, IndexError):
        await callback.message.answer('❌ Ошибка: неверный формат данных')
        return

    # ✅ БЕРЁМ ВСЁ ИЗ STATE!
    data = await state.get_data()
    username = data.get('new_username') or data.get('selected_username') or ''
    name = data.get('pending_name') or data.get('new_name') or data.get('selected_name') or 'Пользователь'

    clan = await get_clan(clan_id)
    if not clan:
        await callback.message.answer('❌ Клан не найден')
        return

    # Удаляем пользователя со всех должностей
    clans = await get_clans()
    for c in clans:
        if len(c) > 2 and c[2] == user_id:
            await remove_clan_leader(c[0])
        if len(c) > 5 and c[5] == user_id:
            await remove_clan_deputy(c[0])

    if role_type == 'leader':
        await update_clan_leader(clan_id, user_id, username, name)
        await callback.message.edit_text(
            f'✅ {name} (@{username}) назначен лидером клана {clan[1]}!\n'
            f'Старая должность автоматически удалена.'
        )
    else:
        await update_clan_deputy(clan_id, user_id, username, name)
        await callback.message.edit_text(
            f'✅ {name} (@{username}) назначен замом клана {clan[1]}!\n'
            f'Старая должность автоматически удалена.'
        )

    await state.clear()
    await callback.message.answer('👥 Управление руководителями\n\nВыберите действие:', reply_markup=manage_roles_menu())

@router.callback_query(F.data == 'role_remove_leader')
async def role_remove_leader(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    await callback.message.edit_text('Выберите клан для удаления лидера:', reply_markup=clan_choice_for_roles())


@router.callback_query(F.data == 'role_remove_deputy')
async def role_remove_deputy(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    await callback.message.edit_text('Выберите клан для удаления зама:', reply_markup=clan_choice_for_roles())


@router.callback_query(F.data.startswith('role_clan_'))
async def role_remove_select_clan(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    clan_id = int(callback.data.split('_')[2])
    clan = await get_clan(clan_id)
    if not clan:
        await callback.message.answer('❌ Клан не найден')
        return

    await callback.message.edit_text(
        f'⚠️ Вы уверены, что хотите удалить руководителя из клана {clan[1]}?',
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text='✅ Да, удалить', callback_data=f'role_confirm_remove_{clan_id}')],
            [InlineKeyboardButton(text='🔙 Назад', callback_data='back_to_admin')],
        ])
    )


@router.callback_query(F.data.startswith('role_confirm_remove_'))
async def role_confirm_remove(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    clan_id = int(callback.data.split('_')[3])
    clan = await get_clan(clan_id)

    if len(clan) > 3 and clan[3]:
        await remove_clan_leader(clan_id)
        await callback.message.edit_text(f'✅ Лидер удалён из клана {clan[1]}')
    elif len(clan) > 6 and clan[6]:
        await remove_clan_deputy(clan_id)
        await callback.message.edit_text(f'✅ Зам удалён из клана {clan[1]}')
    else:
        await callback.message.edit_text(f'❌ В клане {clan[1]} нет руководителей для удаления.')

    await callback.message.answer('👥 Управление руководителями\n\nВыберите действие:', reply_markup=manage_roles_menu())


@router.callback_query(F.data == 'role_list')
async def role_list(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    clans = await get_clans()
    text = '👥 ТЕКУЩИЕ РУКОВОДИТЕЛИ:\n\n'
    emojis = {1: '🔴', 2: '🟡', 3: '🟢', 4: '🟣', 5: '🟠'}

    for clan in clans:
        if len(clan) >= 9:
            clan_id = clan[0]
            name = clan[1]
            leader_id = clan[2] if len(clan) > 2 else None
            leader_username = clan[3] if len(clan) > 3 else None
            leader_name = clan[4] if len(clan) > 4 else None
            deputy_id = clan[5] if len(clan) > 5 else None
            deputy_username = clan[6] if len(clan) > 6 else None
            deputy_name = clan[7] if len(clan) > 7 else None
        else:
            continue
            
        emoji = emojis.get(clan_id, '🔵')

        text += f'{emoji} {name}:\n'
        text += f'   👑 Лидер: {leader_name if leader_name else "❌ не назначен"}'
        if leader_username:
            text += f' (@{leader_username})'
        if leader_id:
            text += f' (ID: {leader_id})'
        text += '\n'

        text += f'   👤 Зам: {deputy_name if deputy_name else "❌ не назначен"}'
        if deputy_username:
            text += f' (@{deputy_username})'
        if deputy_id:
            text += f' (ID: {deputy_id})'
        text += '\n\n'

    await callback.message.edit_text(text, reply_markup=back_button('back_to_admin'))


# ============================================================
# 🗑 ОЧИСТИТЬ ТЕСТОВЫЕ ЗАЯВКИ
# ============================================================

@router.callback_query(F.data == 'admin_clear_test')
async def admin_clear_test(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    count = 0
    url = os.getenv('DATABASE_URL')
    if url:
        conn = await asyncpg.connect(url)
        try:
            count = await conn.fetchval("SELECT COUNT(*) FROM applications WHERE username = 'test_user'")
        finally:
            await conn.close()
    else:
        async with aiosqlite.connect('klan_kaif.db') as db:
            cursor = await db.execute("SELECT COUNT(*) FROM applications WHERE username = 'test_user'")
            result = await cursor.fetchone()
            count = result[0] if result else 0

    if count == 0:
        await callback.message.edit_text(
            '🧪 Нет тестовых заявок для удаления.',
            reply_markup=admin_menu()
        )
        return

    await callback.message.edit_text(
        f'⚠️ ВЫ УВЕРЕНЫ?\n\n'
        f'Будут удалены ВСЕ тестовые заявки (с пометкой "ТЕСТ").\n'
        f'Найдено: {count} тестовых заявок.\n'
        f'Обычные заявки останутся нетронутыми.\n\n'
        f'Это действие нельзя отменить!',
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text='✅ Да, удалить все', callback_data='confirm_clear_test')],
            [InlineKeyboardButton(text='❌ Отмена', callback_data='back_to_admin')],
        ])
    )


@router.callback_query(F.data == 'confirm_clear_test')
async def confirm_clear_test(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    print("🗑 УДАЛЯЕМ ТЕСТОВЫЕ ЗАЯВКИ...")
    await clear_test_applications()
    print("✅ ГОТОВО!")

    await callback.message.edit_text(
        '✅ Все тестовые заявки удалены!\n\n'
        'Обычные заявки остались нетронутыми.',
        reply_markup=admin_menu()
    )


# ============================================================
# 🔄 УПРАВЛЕНИЕ СТАТУСОМ КЛАНОВ (АДМИН)
# ============================================================

@router.callback_query(F.data == 'admin_clan_status')
async def admin_clan_status(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    
    await callback.message.edit_text(
        '🔧 УПРАВЛЕНИЕ СТАТУСОМ КЛАНОВ\n\n'
        'Выберите клан для включения/отключения приёма заявок.\n'
        '❌ Выключенный клан не будет отображаться у кандидатов.',
        reply_markup=admin_clan_status_menu()
    )


@router.callback_query(F.data.startswith('admin_clan_status_'))
async def admin_clan_status_detail(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    
    clan_id = int(callback.data.split('_')[3])
    clan = await get_clan(clan_id)
    if not clan:
        await callback.message.answer('❌ Клан не найден')
        return
    
    is_active = await get_clan_active_status(clan_id)
    status_text = "✅ ВКЛЮЧЁН" if is_active else "❌ ВЫКЛЮЧЁН"
    
    emojis = {1: '🔴', 2: '🟡', 3: '🟢', 4: '🟣', 5: '🟠'}
    emoji = emojis.get(clan_id, '🔵')
    
    await callback.message.edit_text(
        f'{emoji} КЛАН {clan[1]}\n\n'
        f'📌 Статус: {status_text}\n\n'
        f'{"🟢 Заявки принимаются" if is_active else "🔴 Заявки НЕ принимаются"}\n\n'
        f'Нажмите кнопку ниже, чтобы изменить статус.',
        reply_markup=clan_toggle_button(clan_id, clan[1], is_active)
    )


@router.callback_query(F.data.startswith('toggle_clan_'))
async def toggle_clan(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    
    clan_id = int(callback.data.split('_')[2])
    clan = await get_clan(clan_id)
    if not clan:
        await callback.message.answer('❌ Клан не найден')
        return
    
    current_status = await get_clan_active_status(clan_id)
    new_status = not current_status
    
    await set_clan_active(clan_id, new_status)
    
    emojis = {1: '🔴', 2: '🟡', 3: '🟢', 4: '🟣', 5: '🟠'}
    emoji = emojis.get(clan_id, '🔵')
    
    await callback.message.edit_text(
        f'{emoji} КЛАН {clan[1]}\n\n'
        f'✅ Статус изменён на: {"🟢 ВКЛЮЧЁН" if new_status else "🔴 ВЫКЛЮЧЁН"}\n\n'
        f'{"✅ Заявки теперь принимаются" if new_status else "❌ Заявки НЕ принимаются"}\n\n'
        f'Нажмите кнопку ниже, чтобы снова изменить статус.',
        reply_markup=clan_toggle_button(clan_id, clan[1], new_status)
    )


# ============================================================
# 🏗️ МЕНЮ УПРАВЛЕНИЯ КЛАНАМИ
# ============================================================

@router.callback_query(F.data == 'admin_clan_management')
async def admin_clan_management(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    
    await callback.message.edit_text(
        '🏗️ УПРАВЛЕНИЕ КЛАНАМИ\n\n'
        'Выберите действие:\n\n'
        '➕ Добавить новый клан\n'
        '🗑 Удалить существующий клан\n'
        '✏️ Редактировать данные клана',
        reply_markup=admin_clan_management_menu()
    )


# ============================================================
# 📤 АДМИН: ЭКСПОРТ EXCEL
# ============================================================

@router.callback_query(F.data == 'admin_export')
async def admin_export(callback: CallbackQuery):
    print("🔍 НАЖАТА КНОПКА CSV!")
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        apps = await get_all_applications()
        print(f"📊 Найдено заявок: {len(apps)}")
        
        if not apps:
            await callback.message.answer('❌ Нет заявок для экспорта')
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Заявки"

        headers = [
            'ID', 'User ID', 'Username', 'Клан', 'Имя', 'Возраст',
            'Ник', 'ID игровой', 'Часовой пояс', 'Скрин прошлый',
            'Скрин текущий', 'Количество фото', 'Статус', 'Дата создания',
            'Кто одобрил (ID)', 'Кто одобрил (Username)', 'Дата одобрения'
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        status_colors = {
            'accepted': PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
            'rejected': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            'pending': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
            'revoked': PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid"),
        }

        all_leaders = {}
        clans = await get_clans()
        for clan in clans:
            if len(clan) > 3:
                leader_id = clan[2] if len(clan) > 2 else None
                leader_username = clan[3] if len(clan) > 3 else None
                deputy_id = clan[5] if len(clan) > 5 else None
                deputy_username = clan[6] if len(clan) > 6 else None
                
                if leader_id:
                    all_leaders[leader_id] = leader_username or str(leader_id)
                if deputy_id:
                    all_leaders[deputy_id] = deputy_username or str(deputy_id)

        for app in apps:
            try:
                if len(app) >= 12:
                    (app_id, user_id, username, clan_name, answers_json,
                     photo_old, photo_new, has_photos, status,
                     created_at, reviewed_by, reviewed_at) = app[:12]
                else:
                    app_id = app[0] if len(app) > 0 else None
                    user_id = app[1] if len(app) > 1 else None
                    username = app[2] if len(app) > 2 else None
                    clan_name = app[3] if len(app) > 3 else None
                    answers_json = app[4] if len(app) > 4 else '{}'
                    photo_old = app[5] if len(app) > 5 else None
                    photo_new = app[6] if len(app) > 6 else None
                    has_photos = app[7] if len(app) > 7 else 0
                    status = app[8] if len(app) > 8 else 'pending'
                    created_at = app[9] if len(app) > 9 else None
                    reviewed_by = app[10] if len(app) > 10 else None
                    reviewed_at = app[11] if len(app) > 11 else None
            except Exception as e:
                print(f"❌ Ошибка распаковки: {e}")
                continue

            answers = json.loads(answers_json) if answers_json else {}

            if isinstance(created_at, datetime):
                created_at_str = created_at.strftime('%d.%m.%Y %H:%M')
            elif created_at:
                created_at_str = str(created_at)[:16]
            else:
                created_at_str = ''

            if isinstance(reviewed_at, datetime):
                reviewed_at_str = reviewed_at.strftime('%d.%m.%Y %H:%M')
            elif reviewed_at:
                reviewed_at_str = str(reviewed_at)[:16]
            else:
                reviewed_at_str = ''

            is_test = username == 'test_user' or 'Тест' in answers.get('name', '')

            status_ru = {
                'pending': '⏳ На рассмотрении',
                'accepted': '✅ Принято',
                'rejected': '❌ Отклонено',
                'revoked': '⚠️ Отозвано'
            }.get(status, status)

            if is_test:
                status_ru = '🧪 ' + status_ru + ' (ТЕСТ)'

            reviewer_username = ''
            if reviewed_by:
                reviewer_username = all_leaders.get(reviewed_by, '')
                if not reviewer_username:
                    try:
                        async with aiosqlite.connect(DB_PATH) as db:
                            async with db.execute(
                                'SELECT username FROM applications WHERE user_id = ? LIMIT 1',
                                (reviewed_by,)
                            ) as cursor:
                                result = await cursor.fetchone()
                                if result:
                                    reviewer_username = result[0]
                    except:
                        pass
                if not reviewer_username:
                    reviewer_username = str(reviewed_by)

            row = [
                app_id, user_id, f'@{username}' if username else '', clan_name or '',
                answers.get('name', ''), answers.get('age', ''),
                answers.get('nickname', ''), answers.get('id', ''),
                answers.get('timezone', ''),
                '✅' if photo_old else '❌',
                '✅' if photo_new else '❌',
                has_photos or 0,
                status_ru,
                created_at_str,
                reviewed_by if reviewed_by else '',
                reviewer_username,
                reviewed_at_str
            ]
            ws.append(row)

            row_num = ws.max_row

            if is_test:
                fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
                font_color = "FFFFFF"
            else:
                if status == 'accepted':
                    fill = status_colors['accepted']
                    font_color = "000000"
                elif status == 'rejected':
                    fill = status_colors['rejected']
                    font_color = "FFFFFF"
                elif status == 'pending':
                    fill = status_colors['pending']
                    font_color = "000000"
                else:
                    fill = status_colors['revoked']
                    font_color = "000000"

            status_cell = ws.cell(row=row_num, column=13)
            status_cell.fill = fill
            status_cell.font = Font(bold=True, color=font_color, size=10)
            status_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = max_length + 3

        ws.freeze_panes = 'A2'

        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from aiogram.types import BufferedInputFile
        await callback.message.answer_document(
            document=BufferedInputFile(output.getvalue(), filename='заявки.xlsx'),
            caption='📊 Все заявки в формате Excel с цветами!\n\n✅ Добавлены колонки: Кто одобрил, Дата одобрения, Количество фото'
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        await callback.message.answer(f'❌ Ошибка при экспорте: {e}')


# ============================================================
# 📊 АДМИН: СТАТИСТИКА
# ============================================================

@router.callback_query(F.data == 'admin_stats')
async def admin_stats(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    stats, by_clan = await get_statistics()
    total, pending, accepted, rejected, revoked = stats

    text = f'📊 СТАТИСТИКА ЗАЯВОК:\n\nВсего: {total}\n⏳ На рассмотрении: {pending}\n✅ Принято: {accepted}\n❌ Отклонено: {rejected}\n⚠️ Отозвано: {revoked}\n\nПо кланам:\n'
    for clan_name, count in by_clan:
        text += f'   {clan_name}: {count} заявок\n'

    await callback.message.edit_text(text, reply_markup=admin_menu())


# ============================================================
# 👥 АДМИН: ЧЁРНЫЙ СПИСОК
# ============================================================

@router.callback_query(F.data == 'admin_blacklist')
async def admin_blacklist(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()

    blacklist = await get_blacklist()
    if not blacklist:
        text = '👥 Чёрный список пуст.'
    else:
        text = '👥 ЧЁРНЫЙ СПИСОК:\n\n'
        for item in blacklist:
            text += f'ID: {item[1]}\nПричина: {item[2]}\nДобавлен: {item[4][:10] if item[4] else "неизвестно"}\n\n'

    await callback.message.edit_text(text, reply_markup=admin_menu())


# ============================================================
# ➕ ДОБАВЛЕНИЕ КЛАНА
# ============================================================

@router.callback_query(F.data == 'admin_add_clan')
async def admin_add_clan_start(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    
    await state.set_state(ClanManagementForm.waiting_new_clan_name)
    await callback.message.edit_text(
        '🏗️ ДОБАВЛЕНИЕ НОВОГО КЛАНА\n\n'
        'Введите НАЗВАНИЕ клана:\n'
        'Например: "MY CLAN" или "LEGENDS"\n\n'
        '⚠️ Название должно быть уникальным.',
        reply_markup=back_button('back_to_admin')
    )


@router.message(ClanManagementForm.waiting_new_clan_name)
async def admin_add_clan_name(message: Message, state: FSMContext):
    name = message.text.strip()
    if not name:
        await message.answer('❌ Название не может быть пустым. Попробуйте снова:')
        return
    
    existing = await get_clan_by_name(name)
    if existing:
        await message.answer(f'❌ Клан "{name}" уже существует! Введите другое название:')
        return
    
    await state.update_data(new_clan_name=name)
    await state.set_state(ClanManagementForm.waiting_new_clan_emoji)
    await message.answer(
        f'✅ Название "{name}" сохранено!\n\n'
        'Введите ЭМОДЗИ для клана:\n'
        'Например: 🔵 или ⚔️ или 🏆\n\n'
        'Или отправьте "пропустить" чтобы оставить 🔵',
        reply_markup=back_button('back_to_admin')
    )


@router.message(ClanManagementForm.waiting_new_clan_emoji)
async def admin_add_clan_emoji(message: Message, state: FSMContext):
    text = message.text.strip()
    if text.lower() == 'пропустить' or text.lower() == 'skip':
        emoji = '🔵'
    else:
        emoji = text
    
    await state.update_data(new_clan_emoji=emoji)
    await state.set_state(ClanManagementForm.waiting_new_clan_leader_id)
    await message.answer(
        f'✅ Эмодзи: {emoji}\n\n'
        'Введите TELEGRAM ID лидера клана (число):\n'
        'Например: 123456789\n\n'
        'Или отправьте "пропустить" чтобы оставить без лидера',
        reply_markup=back_button('back_to_admin')
    )


@router.message(ClanManagementForm.waiting_new_clan_leader_id)
async def admin_add_clan_leader_id(message: Message, state: FSMContext):
    text = message.text.strip()
    leader_id = None
    if text.lower() not in ['пропустить', 'skip']:
        try:
            leader_id = int(text)
        except:
            await message.answer('❌ Введите число или "пропустить":')
            return
    
    await state.update_data(new_leader_id=leader_id)
    await state.set_state(ClanManagementForm.waiting_new_clan_leader_username)
    await message.answer(
        f'✅ ID лидера: {leader_id if leader_id else "не указан"}\n\n'
        'Введите USERNAME лидера (без @):\n'
        'Например: username\n\n'
        'Или отправьте "пропустить"',
        reply_markup=back_button('back_to_admin')
    )


@router.message(ClanManagementForm.waiting_new_clan_leader_username)
async def admin_add_clan_leader_username(message: Message, state: FSMContext):
    text = message.text.strip()
    username = None if text.lower() in ['пропустить', 'skip'] else text.replace('@', '')
    
    await state.update_data(new_leader_username=username)
    await state.set_state(ClanManagementForm.waiting_new_clan_leader_name)
    await message.answer(
        f'✅ Username лидера: {username if username else "не указан"}\n\n'
        'Введите ИМЯ лидера:\n'
        'Например: Антон\n\n'
        'Или отправьте "пропустить"',
        reply_markup=back_button('back_to_admin')
    )


@router.message(ClanManagementForm.waiting_new_clan_leader_name)
async def admin_add_clan_leader_name(message: Message, state: FSMContext):
    text = message.text.strip()
    name = None if text.lower() in ['пропустить', 'skip'] else text
    
    await state.update_data(new_leader_name=name)
    await state.set_state(ClanManagementForm.waiting_new_clan_deputy_id)
    await message.answer(
        f'✅ Имя лидера: {name if name else "не указан"}\n\n'
        'Введите TELEGRAM ID зама клана (число):\n'
        'Или отправьте "пропустить" чтобы оставить без зама',
        reply_markup=back_button('back_to_admin')
    )


@router.message(ClanManagementForm.waiting_new_clan_deputy_id)
async def admin_add_clan_deputy_id(message: Message, state: FSMContext):
    text = message.text.strip()
    deputy_id = None
    if text.lower() not in ['пропустить', 'skip']:
        try:
            deputy_id = int(text)
        except:
            await message.answer('❌ Введите число или "пропустить":')
            return
    
    await state.update_data(new_deputy_id=deputy_id)
    await state.set_state(ClanManagementForm.waiting_new_clan_deputy_username)
    await message.answer(
        f'✅ ID зама: {deputy_id if deputy_id else "не указан"}\n\n'
        'Введите USERNAME зама (без @):\n'
        'Или отправьте "пропустить"',
        reply_markup=back_button('back_to_admin')
    )


@router.message(ClanManagementForm.waiting_new_clan_deputy_username)
async def admin_add_clan_deputy_username(message: Message, state: FSMContext):
    text = message.text.strip()
    username = None if text.lower() in ['пропустить', 'skip'] else text.replace('@', '')
    
    await state.update_data(new_deputy_username=username)
    await state.set_state(ClanManagementForm.waiting_new_clan_deputy_name)
    await message.answer(
        f'✅ Username зама: {username if username else "не указан"}\n\n'
        'Введите ИМЯ зама:\n'
        'Или отправьте "пропустить"',
        reply_markup=back_button('back_to_admin')
    )


@router.message(ClanManagementForm.waiting_new_clan_deputy_name)
async def admin_add_clan_deputy_name(message: Message, state: FSMContext):
    text = message.text.strip()
    name = None if text.lower() in ['пропустить', 'skip'] else text
    
    data = await state.get_data()
    
    try:
        clan_id = await add_clan(
            name=data['new_clan_name'],
            emoji=data.get('new_clan_emoji', '🔵'),
            leader_id=data.get('new_leader_id'),
            leader_username=data.get('new_leader_username'),
            leader_name=data.get('new_leader_name'),
            deputy_id=data.get('new_deputy_id'),
            deputy_username=data.get('new_deputy_username'),
            deputy_name=name
        )
        
        emoji = data.get('new_clan_emoji', '🔵')
        
        await state.clear()
        await message.answer(
            f'✅ Клан "{data["new_clan_name"]}" {emoji} создан!\n\n'
            f'📋 Данные:\n'
            f'ID: {clan_id}\n'
            f'Название: {data["new_clan_name"]}\n'
            f'Эмодзи: {emoji}\n'
            f'Лидер: {data.get("new_leader_name", "не назначен")}\n'
            f'Зам: {name if name else "не назначен"}\n\n'
            f'Теперь вы можете настроить ссылку на чат через кнопку "Связаться" в заявке.',
            reply_markup=admin_clan_management_menu()
        )
        
    except Exception as e:
        await message.answer(f'❌ Ошибка при создании клана: {e}', reply_markup=admin_clan_management_menu())


# ============================================================
# 🗑 УДАЛЕНИЕ КЛАНА
# ============================================================

@router.callback_query(F.data == 'admin_delete_clan')
async def admin_delete_clan_start(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    
    clans = await get_clans()
    buttons = []
    emojis = {1: '🔴', 2: '🟡', 3: '🟢', 4: '🟣', 5: '🟠'}
    
    for clan in clans:
        clan_id = clan[0]
        name = clan[1]
        emoji = emojis.get(clan_id, '🔵')
        buttons.append([InlineKeyboardButton(
            text=f'{emoji} {name}',
            callback_data=f'admin_delete_clan_confirm_{clan_id}'
        )])
    
    buttons.append([InlineKeyboardButton(text='🔙 Назад', callback_data='back_to_clan_management')])
    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
    
    await callback.message.edit_text(
        '🗑 УДАЛЕНИЕ КЛАНА\n\n'
        'Выберите клан для удаления:\n'
        '⚠️ ВНИМАНИЕ: ВСЕ ЗАЯВКИ ЭТОГО КЛАНА ТОЖЕ БУДУТ УДАЛЕНЫ!\n'
        'Это действие нельзя отменить!',
        reply_markup=keyboard
    )


@router.callback_query(F.data.startswith('admin_delete_clan_confirm_'))
async def admin_delete_clan_confirm(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    
    clan_id = int(callback.data.split('_')[4])
    clan = await get_clan(clan_id)
    if not clan:
        await callback.message.answer('❌ Клан не найден')
        return
    
    apps = await get_clan_applications(clan_id)
    count = len(apps)
    
    await callback.message.edit_text(
        f'⚠️ ПОДТВЕРЖДЕНИЕ УДАЛЕНИЯ\n\n'
        f'Вы уверены, что хотите удалить клан "{clan[1]}"?\n\n'
        f'📊 Заявок в клане: {count}\n\n'
        f'❌ ВСЕ заявки будут удалены безвозвратно!\n'
        f'Это действие нельзя отменить!',
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text='✅ ДА, УДАЛИТЬ', callback_data=f'admin_delete_clan_execute_{clan_id}')],
            [InlineKeyboardButton(text='❌ ОТМЕНА', callback_data='admin_delete_clan')],
        ])
    )


@router.callback_query(F.data.startswith('admin_delete_clan_execute_'))
async def admin_delete_clan_execute(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    
    clan_id = int(callback.data.split('_')[4])
    clan = await get_clan(clan_id)
    if not clan:
        await callback.message.answer('❌ Клан не найден')
        return
    
    clan_name = clan[1]
    
    try:
        await delete_clan(clan_id)
        await callback.message.edit_text(
            f'✅ Клан "{clan_name}" УДАЛЁН!\n\n'
            f'Все связанные заявки и ссылки также удалены.',
            reply_markup=admin_clan_management_menu()
        )
    except Exception as e:
        await callback.message.answer(f'❌ Ошибка при удалении: {e}', reply_markup=admin_clan_management_menu())


# ============================================================
# ✏️ РЕДАКТИРОВАНИЕ КЛАНА
# ============================================================

@router.callback_query(F.data == 'admin_edit_clan')
async def admin_edit_clan_start(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    
    clans = await get_clans()
    buttons = []
    emojis = {1: '🔴', 2: '🟡', 3: '🟢', 4: '🟣', 5: '🟠'}
    
    for clan in clans:
        clan_id = clan[0]
        name = clan[1]
        emoji = emojis.get(clan_id, '🔵')
        buttons.append([InlineKeyboardButton(
            text=f'{emoji} {name}',
            callback_data=f'admin_edit_clan_select_{clan_id}'
        )])
    
    buttons.append([InlineKeyboardButton(text='🔙 Назад', callback_data='back_to_clan_management')])
    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
    
    await callback.message.edit_text(
        '✏️ РЕДАКТИРОВАНИЕ КЛАНА\n\n'
        'Выберите клан для редактирования:',
        reply_markup=keyboard
    )


@router.callback_query(F.data.startswith('admin_edit_clan_select_'))
async def admin_edit_clan_select(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    
    clan_id = int(callback.data.split('_')[4])
    clan = await get_clan(clan_id)
    if not clan:
        await callback.message.answer('❌ Клан не найден')
        return
    
    if len(clan) >= 11:
        clan_id, name, emoji, leader_id, leader_username, leader_name, deputy_id, deputy_username, deputy_name, is_active, created_at = clan
    else:
        clan_id = clan[0]
        name = clan[1]
        emoji = clan[2] if len(clan) > 2 else '🔵'
        leader_id = clan[3] if len(clan) > 3 else None
        leader_username = clan[4] if len(clan) > 4 else None
        leader_name = clan[5] if len(clan) > 5 else None
        deputy_id = clan[6] if len(clan) > 6 else None
        deputy_username = clan[7] if len(clan) > 7 else None
        deputy_name = clan[8] if len(clan) > 8 else None
        is_active = clan[9] if len(clan) > 9 else True
        created_at = clan[10] if len(clan) > 10 else None
    
    await callback.message.edit_text(
        f'✏️ РЕДАКТИРОВАНИЕ КЛАНА {emoji} {name}\n\n'
        f'📋 ТЕКУЩИЕ ДАННЫЕ:\n'
        f'Название: {name}\n'
        f'Эмодзи: {emoji}\n'
        f'Лидер: {leader_name or "❌"} (@{leader_username or "❌"}) ID: {leader_id or "❌"}\n'
        f'Зам: {deputy_name or "❌"} (@{deputy_username or "❌"}) ID: {deputy_id or "❌"}\n\n'
        f'Выберите что изменить:',
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text='📝 Изменить название', callback_data=f'edit_clan_field_name_{clan_id}')],
            [InlineKeyboardButton(text='🔄 Изменить эмодзи', callback_data=f'edit_clan_field_emoji_{clan_id}')],
            [InlineKeyboardButton(text='👑 Изменить лидера', callback_data=f'edit_clan_field_leader_{clan_id}')],
            [InlineKeyboardButton(text='👤 Изменить зама', callback_data=f'edit_clan_field_deputy_{clan_id}')],
            [InlineKeyboardButton(text='🔙 Назад', callback_data='admin_edit_clan')],
        ])
    )


@router.callback_query(F.data.startswith('edit_clan_field_'))
async def admin_edit_clan_field(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    
    parts = callback.data.split('_')
    field = parts[3]
    clan_id = int(parts[4])
    
    await state.update_data(edit_clan_id=clan_id, edit_field=field)
    
    field_names = {
        'name': 'НАЗВАНИЕ',
        'emoji': 'ЭМОДЗИ',
        'leader': 'ЛИДЕРА (введите ID, username, имя через запятую)',
        'deputy': 'ЗАМА (введите ID, username, имя через запятую)'
    }
    
    examples = {
        'name': 'Пример: MY NEW CLAN',
        'emoji': 'Пример: ⚔️',
        'leader': 'Пример: 123456789, username, Имя',
        'deputy': 'Пример: 987654321, deputy_username, Имя Зама'
    }
    
    await callback.message.edit_text(
        f'✏️ ИЗМЕНЕНИЕ {field_names.get(field, field)}\n\n'
        f'Введите новые данные:\n{examples.get(field, "")}\n\n'
        f'Или отправьте "пропустить" чтобы оставить как есть.',
        reply_markup=back_button('admin_edit_clan')
    )
    
    await state.set_state(ClanManagementForm.waiting_edit_clan_field)


@router.message(ClanManagementForm.waiting_edit_clan_field)
async def admin_edit_clan_field_value(message: Message, state: FSMContext):
    data = await state.get_data()
    clan_id = data.get('edit_clan_id')
    field = data.get('edit_field')
    
    text = message.text.strip()
    
    if text.lower() in ['пропустить', 'skip']:
        await state.clear()
        await message.answer('✅ Изменение отменено.', reply_markup=admin_clan_management_menu())
        return
    
    try:
        if field == 'name':
            existing = await get_clan_by_name(text)
            if existing and existing[0] != clan_id:
                await message.answer(f'❌ Клан "{text}" уже существует! Введите другое название:')
                return
            await update_clan(clan_id, name=text)
            await message.answer(f'✅ Название изменено на "{text}"!')
            
        elif field == 'emoji':
            await update_clan(clan_id, emoji=text)
            await message.answer(f'✅ Эмодзи изменён на {text}!')
            
        elif field == 'leader':
            parts = [p.strip() for p in text.split(',')]
            leader_id = int(parts[0]) if len(parts) > 0 and parts[0].isdigit() else None
            leader_username = parts[1] if len(parts) > 1 and parts[1] != 'пропустить' else None
            leader_name = parts[2] if len(parts) > 2 and parts[2] != 'пропустить' else None
            
            await update_clan(clan_id, leader_id=leader_id, leader_username=leader_username, leader_name=leader_name)
            await message.answer(f'✅ Лидер обновлён!\nID: {leader_id}\nUsername: {leader_username}\nИмя: {leader_name}')
            
        elif field == 'deputy':
            parts = [p.strip() for p in text.split(',')]
            deputy_id = int(parts[0]) if len(parts) > 0 and parts[0].isdigit() else None
            deputy_username = parts[1] if len(parts) > 1 and parts[1] != 'пропустить' else None
            deputy_name = parts[2] if len(parts) > 2 and parts[2] != 'пропустить' else None
            
            await update_clan(clan_id, deputy_id=deputy_id, deputy_username=deputy_username, deputy_name=deputy_name)
            await message.answer(f'✅ Зам обновлён!\nID: {deputy_id}\nUsername: {deputy_username}\nИмя: {deputy_name}')
        
        await state.clear()
        await message.answer('📋 Что дальше?', reply_markup=admin_clan_management_menu())
        
    except Exception as e:
        await message.answer(f'❌ Ошибка: {e}. Попробуйте снова:', reply_markup=back_button('admin_edit_clan'))


# ============================================================
# 🔙 НАЗАД К УПРАВЛЕНИЮ РУКОВОДИТЕЛЯМИ
# ============================================================

@router.callback_query(F.data == 'back_to_roles')
async def back_to_roles(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    await state.clear()
    await callback.message.edit_text(
        '👥 Управление руководителями\n\n'
        'Здесь вы можете назначить или удалить лидера/зама для любого клана.',
        reply_markup=manage_roles_menu()
    )


# ============================================================
# 🔙 НАЗАД К ВЫБОРУ СПОСОБА НАЗНАЧЕНИЯ
# ============================================================

@router.callback_query(F.data == 'select_existing_choice')
async def select_existing_choice(callback: CallbackQuery, state: FSMContext):
    """Назад к выбору способа назначения (из выбора клана)"""
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    
    data = await state.get_data()
    role_type = data.get('role_type', 'leader')
    
    if role_type == 'leader':
        await callback.message.edit_text(
            '👥 Назначение лидера\n\n'
            'Выберите действие:',
            reply_markup=assign_choice_menu()
        )
    else:
        await callback.message.edit_text(
            '👥 Назначение зама\n\n'
            'Выберите действие:',
            reply_markup=assign_choice_menu()
        )


# ============================================================
# 🔙 НАЗАД К ВЫБОРУ СПОСОБА НАЗНАЧЕНИЯ (ИЗ СПИСКА РУКОВОДИТЕЛЕЙ)
# ============================================================

@router.callback_query(F.data == 'assign_choice_menu')
async def assign_choice_menu_back(callback: CallbackQuery, state: FSMContext):
    """Назад к выбору способа назначения (из списка руководителей)"""
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    
    data = await state.get_data()
    role_type = data.get('role_type', 'leader')
    
    if role_type == 'leader':
        await callback.message.edit_text(
            '👥 Назначение лидера\n\n'
            'Выберите действие:',
            reply_markup=assign_choice_menu()
        )
    else:
        await callback.message.edit_text(
            '👥 Назначение зама\n\n'
            'Выберите действие:',
            reply_markup=assign_choice_menu()
        )


# ============================================================
# 🔙 НАЗАД К УПРАВЛЕНИЮ РУКОВОДИТЕЛЯМИ (ИЗ УДАЛЕНИЯ)
# ============================================================

@router.callback_query(F.data == 'back_to_manage_roles')
async def back_to_manage_roles(callback: CallbackQuery, state: FSMContext):
    """Назад в управление руководителями"""
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer('⛔ Нет прав')
        return
    await callback.answer()
    await state.clear()
    await callback.message.edit_text(
        '👥 Управление руководителями\n\n'
        'Здесь вы можете назначить или удалить лидера/зама для любого клана.',
        reply_markup=manage_roles_menu()
    )
